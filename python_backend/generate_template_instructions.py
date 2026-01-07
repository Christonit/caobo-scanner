"""
Generate instructions for server.py based on template structure
Based on template analysis: columns are Documento, Tipo de Suplidor, Tipo de Gasto, Descripcion, Fecha, Monto en Servicios
"""
import json
from pathlib import Path

BASE_DIR = Path(__file__).parent
OUTPUT_FILE = BASE_DIR / "template_instructions.json"

# Based on the template structure observed:
# Sheet: "Listado de Gastos"
# Columns: Documento, Tipo de Suplidor, Tipo de Gasto, Descripcion, Fecha, Monto en Servicios
# Header row: 1 (yellow background)
# Auto-filter: Enabled

instructions = {
    "template_info": {
        "file": "template.xls",
        "main_sheet": "Listado de Gastos",
        "header_row": 1,
        "data_start_row": 2,
        "auto_filter_enabled": True,
        "header_background_color": "FFFFFF00",  # Yellow
    },
    "column_structure": {
        "columns": [
            {"index": 1, "name": "Documento", "letter": "A"},
            {"index": 2, "name": "Tipo de Suplidor", "letter": "B"},
            {"index": 3, "name": "Tipo de Gasto", "letter": "C"},
            {"index": 4, "name": "Descripcion", "letter": "D"},
            {"index": 5, "name": "Fecha", "letter": "E"},
            {"index": 6, "name": "Monto en Servicios", "letter": "F"},
        ]
    },
    "data_mapping": {
        "description": {
            "template_column": 4,
            "template_column_letter": "D",
            "template_header": "Descripcion",
            "source_field": "description",  # From Gemini extraction
            "fallback": "vendor"  # Use vendor if description not available
        },
        "date": {
            "template_column": 5,
            "template_column_letter": "E",
            "template_header": "Fecha",
            "source_field": "date"
        },
        "total": {
            "template_column": 6,
            "template_column_letter": "F",
            "template_header": "Monto en Servicios",
            "source_field": "total"
        },
        "documento": {
            "template_column": 1,
            "template_column_letter": "A",
            "template_header": "Documento",
            "source_field": "filename"  # Use filename as documento
        },
        # Note: Tipo de Suplidor (B) and Tipo de Gasto (C) are not mapped
        # These might need to be set to default values or left empty
    },
    "code_instructions": [
        "1. Always load template from template.xls (convert to xlsx first if needed)",
        "2. Use sheet 'Listado de Gastos' (not active sheet)",
        "3. Find first empty row starting from row 2 (row 1 is header)",
        "4. Map data as follows:",
        "   - Column A (Documento): Use filename from data",
        "   - Column B (Tipo de Suplidor): Leave empty or set default",
        "   - Column C (Tipo de Gasto): Leave empty or set default",
        "   - Column D (Descripcion): Use vendor/description from data",
        "   - Column E (Fecha): Use date from data",
        "   - Column F (Monto en Servicios): Use total from data",
        "5. Preserve header row formatting (yellow background, bold)",
        "6. Apply auto-filter to header row (row 1)",
        "7. Preserve all column widths and formatting",
        "8. Save to output.xlsx"
    ],
    "python_code_snippet": '''
# In populate_excel_template function:
# 1. Load template and get specific sheet
wb = openpyxl.load_workbook(template_path)
ws = wb["Listado de Gastos"]  # Use specific sheet name, not active

# 2. Find first empty row (start from row 2)
row = 2
while row <= ws.max_row and ws.cell(row=row, column=1).value is not None:
    row += 1

# 3. Map data to template columns
ws.cell(row=row, column=1, value=data.get("filename", ""))  # Documento
# Column 2 (Tipo de Suplidor) - leave empty or set default
# Column 3 (Tipo de Gasto) - leave empty or set default
ws.cell(row=row, column=4, value=data.get("vendor", data.get("description", "")))  # Descripcion
ws.cell(row=row, column=5, value=data.get("date", ""))  # Fecha
ws.cell(row=row, column=6, value=data.get("total", 0))  # Monto en Servicios

# 4. Apply header formatting if not already present
from openpyxl.styles import Font, PatternFill
yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
bold_font = Font(bold=True)

for col in range(1, 7):  # Columns A-F
    header_cell = ws.cell(row=1, column=col)
    if not header_cell.fill.start_color or header_cell.fill.start_color.rgb != "FFFFFF00":
        header_cell.fill = yellow_fill
    if not header_cell.font.bold:
        header_cell.font = bold_font

# 5. Apply auto-filter
ws.auto_filter.ref = f"A1:F{ws.max_row}"

# 6. Save
wb.save(OUTPUT_FILE)
'''
}

# Save instructions
with open(OUTPUT_FILE, 'w', encoding='utf-8') as f:
    json.dump(instructions, f, indent=2, ensure_ascii=False)

print("=" * 80)
print("TEMPLATE INSTRUCTIONS GENERATED")
print("=" * 80)
print(f"Saved to: {OUTPUT_FILE}")
print("\nKey Points:")
print(f"  - Main sheet: {instructions['template_info']['main_sheet']}")
print(f"  - Header row: {instructions['template_info']['header_row']}")
print(f"  - Columns: {', '.join([c['name'] for c in instructions['column_structure']['columns']])}")
print("\nData Mapping:")
for field, mapping in instructions['data_mapping'].items():
    print(f"  {field}: {mapping['source_field']} -> Column {mapping['template_column']} ({mapping['template_header']})")
print("\n" + "=" * 80)

