"""
FastAPI server for processing receipts/invoices with Gemini AI — gastos (expenses) flow.

Suplidores (suppliers) routes live in suplidores_server.py and are mounted
via app.include_router() at the bottom of this file.
"""
import asyncio
import token
from fastapi import FastAPI, UploadFile, File, HTTPException, Body, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
import hashlib
import json
import os
from pathlib import Path
import re
from typing import Any, List, Optional, Tuple
from datetime import date, datetime, time
import google.generativeai as genai
from dotenv import load_dotenv

from shared_utils import (
    GEMINI_MAX_CONCURRENT,
    PDF_MAX_PAGES,
    PDF_RENDER_DPI,
    _is_retryable_error,
    _parse_json_loose,
    gemini_semaphore,
    generate_inference_content,
    get_thinking_level_models,
    load_file_as_images,
    render_pdf_to_images,
    resolve_inference_model,
)
from token_usage import (
    merge_usage_records,
    record_usage_from_response,
    resolve_thinking_level,
)
from xls_template import fill_gastos_xls_template


# Allowed catalog values from the Carga Masiva "Nomencladores" sheet.
TIPO_DE_SUPLIDOR_OPTIONS = [
    "Gasto Formal",
    "Gasto Informal",
    "Genérico",
    "Gasto Menor",
    "Pagos al exterior",
    "Norma 07-2007",
    "DGA",
    "Decreto 139-98",
]

TIPO_DE_GASTO_OPTIONS = [
    "01-Gasto de personal",
    "02-Gastos por trabajos, servicios y suministros",
    "03-Arrendamientos",
    "04-Gastos de activo fijo",
    "05-Gastos de representación",
    "06-Otras deducciones administrativas",
    "07-Gastos financieros",
    "08-Gastos extraordinarios",
    "09-Compras y gastos que forman gastos de la venta",
    "10-Adquisicion de activos",
    "11-Gastos de seguros",
]

# Canonical "Moneda" values: capitalized (first letter only), not upper/title
# cased - "Peso dominicano" keeps "dominicano" lowercase, "Dólar americano"
# uses full country qualifier to match the Excel template label.
MONEDA_OPTIONS = ["Peso dominicano", "Dolar Americano", "Euros"]
MONEDA_ALIASES = {
    "dop": "Peso dominicano",
    "rd$": "Peso dominicano",
    "peso": "Peso dominicano",
    "pesos": "Peso dominicano",
    "peso dominicano": "Peso dominicano",
    "pesos dominicanos": "Peso dominicano",
    "usd": "Dolar Americano",
    "us$": "Dolar Americano",
    "$": "Dolar Americano",
    "dolar": "Dolar Americano",
    "dolares": "Dolar Americano",
    "dólar": "Dolar Americano",
    "dólares": "Dolar Americano",
    "dollar": "Dolar Americano",
    "dollars": "Dolar Americano",
    "dólar americano": "Dolar Americano",
    "dolar americano": "Dolar Americano",
    "eur": "Euros",
    "€": "Euros",
    "euro": "Euros",
    "euros": "Euros",
}

# NCF series that require "NCF Afectado" (credit/debit notes).
NCF_AFECTADO_REQUIRED_PREFIXES = ("B03", "B04")


BASE_DIR = Path(__file__).parent
DATA_DIR = BASE_DIR

env_path = BASE_DIR / ".env"
if env_path.exists():
    load_dotenv(env_path)
    print(f"[INFO] Loaded .env from: {env_path}")
else:
    load_dotenv()

GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
if not GEMINI_API_KEY:
    print("[WARNING] GEMINI_API_KEY not set. AI features will not work.")
    print("[WARNING] Create a .env file with GEMINI_API_KEY=your_key")
else:
    genai.configure(api_key=GEMINI_API_KEY)

app = FastAPI(title="Receipt Processing API")

# CORS: allow the Nuxt frontend (and any local dev origin) to call the API.
# In production, set ALLOWED_ORIGINS to a comma-separated list of trusted URLs.
allowed_origins_env = os.getenv("ALLOWED_ORIGINS", "*")
allowed_origins = (
    ["*"] if allowed_origins_env.strip() == "*"
    else [o.strip() for o in allowed_origins_env.split(",") if o.strip()]
)
app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

HISTORY_FILE = DATA_DIR / "history.json"
# Authoritative Carga Masiva Gastos template. We FILL this file (preserving
# its dropdowns / data-validation named ranges / Nomencladores sheet) instead
# of recreating it, because the destination system rejects a rebuilt workbook.
GASTOS_TEMPLATE_XLS_SOURCE = BASE_DIR / "assets/templates/template-gastos.xls"
# Internal working copy stays .xlsx (openpyxl); downloadable export is .xls
# because destination systems (e.g. Carga Masiva) reject .xlsx uploads.
GASTOS_OUTPUT_FILE = DATA_DIR / "output.xlsx"
GASTOS_OUTPUT_XLS = DATA_DIR / "output.xls"

print(f"[INFO] Base directory: {BASE_DIR}")
print(f"[INFO] Data directory: {DATA_DIR}")
print(f"[INFO] Gastos template: {GASTOS_TEMPLATE_XLS_SOURCE}")

# Ensure history file exists
if not HISTORY_FILE.exists():
    with open(HISTORY_FILE, 'w') as f:
        json.dump([], f)

# Default models when the client does not send a thinking-level `model`.
# Rapido → 3.1-flash-lite, Moderado → 3.5-flash-lite, Profundo → 3.6-flash.
_LEVEL_MODELS = get_thinking_level_models()
INDIVIDUAL_MODEL = _LEVEL_MODELS["rapido"]
BATCH_MODEL = _LEVEL_MODELS["moderado"]


def list_available_gemini_models(*, generate_content_only: bool = True) -> list[dict]:
    """
    Fetch the models available to this Gemini API key via ModelService.ListModels.

    Returns a list of dicts with name, display_name, description, and
    supported_generation_methods. By default only includes models that support
    generateContent (what this server uses for receipt scanning).
    """
    if not GEMINI_API_KEY:
        raise RuntimeError("GEMINI_API_KEY is not configured")

    models: list[dict] = []
    for model in genai.list_models():
        methods = list(model.supported_generation_methods or [])
        if generate_content_only and "generateContent" not in methods:
            continue
        # API returns names like "models/gemini-2.0-flash"; expose both forms.
        full_name = model.name or ""
        short_name = full_name.split("/", 1)[-1] if full_name else ""
        models.append({
            "name": short_name,
            "full_name": full_name,
            "display_name": getattr(model, "display_name", None) or short_name,
            "description": getattr(model, "description", None) or "",
            "supported_generation_methods": methods,
            "input_token_limit": getattr(model, "input_token_limit", None),
            "output_token_limit": getattr(model, "output_token_limit", None),
        })

    models.sort(key=lambda m: m["name"])
    return models


def calculate_file_hash(file_content: bytes) -> str:
    """Calculate MD5 hash of file content"""
    return hashlib.md5(file_content).hexdigest()


def load_history() -> list:
    """Load processing history from JSON file"""
    try:
        with open(HISTORY_FILE, 'r') as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return []


def save_history(history: list):
    """Save processing history to JSON file"""
    with open(HISTORY_FILE, 'w') as f:
        json.dump(history, f, indent=2)


def check_duplicate(file_hash: str) -> bool:
    """Check if file hash exists in history"""
    history = load_history()
    return any(entry.get('hash') == file_hash for entry in history)


def _new_dedupe_state() -> Tuple[set, dict, set, dict]:
    """
    Fresh, empty lookup structures used to detect duplicate receipts WITHIN
    a single upload/batch request only (e.g. the same page appearing twice
    in a multi-page PDF, or the same receipt photographed/selected more than
    once in one "Process All" run).

    Returns (seen_hashes, hash_to_filename, seen_keys, key_to_filename).
    Callers add to these in-place as they walk the current request's files.

    Deliberately NOT seeded from history.json: persisting duplicate
    detection across sessions meant re-scanning the exact same document
    after reloading the page (e.g. to retry/verify) would be permanently
    flagged as "duplicate" and silently skipped, which is confusing and not
    what users expect - duplicate detection should only guard against
    accidentally including the same receipt twice in one export, not block
    intentional re-processing later.
    """
    return set(), {}, set(), {}


def convert_xlsx_to_xls(xlsx_path: Path, xls_path: Path) -> Path:
    """Convert an .xlsx workbook to Excel 97-2003 .xls (BIFF8)."""
    import xlwt
    from openpyxl.utils import get_column_letter

    wb_in = openpyxl.load_workbook(xlsx_path, data_only=True)
    wb_out = xlwt.Workbook(encoding="utf-8")

    date_style = xlwt.XFStyle()
    date_style.num_format_str = "D/MM/YYYY"

    for sheet_name in wb_in.sheetnames:
        ws_in = wb_in[sheet_name]
        ws_out = wb_out.add_sheet(sheet_name[:31])

        max_row = ws_in.max_row or 1
        max_col = min(ws_in.max_column or 1, 256)

        skip_cells = set()
        for merged_range in ws_in.merged_cells.ranges:
            min_r, min_c = merged_range.min_row, merged_range.min_col
            for r in range(merged_range.min_row, merged_range.max_row + 1):
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    if (r, c) != (min_r, min_c):
                        skip_cells.add((r, c))

        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                if (row_idx, col_idx) in skip_cells:
                    continue
                value = ws_in.cell(row=row_idx, column=col_idx).value
                if value is None:
                    continue
                r, c = row_idx - 1, col_idx - 1
                if isinstance(value, datetime):
                    ws_out.write(r, c, value, date_style)
                elif isinstance(value, date):
                    ws_out.write(r, c, datetime.combine(value, time.min), date_style)
                elif isinstance(value, (int, float, bool)):
                    ws_out.write(r, c, value)
                else:
                    ws_out.write(r, c, str(value))

        for col_idx in range(1, max_col + 1):
            dim = ws_in.column_dimensions.get(get_column_letter(col_idx))
            if dim and dim.width:
                ws_out.col(col_idx - 1).width = int(min(float(dim.width), 60) * 256)

        for merged_range in ws_in.merged_cells.ranges:
            if merged_range.max_col > 256:
                continue
            ws_out.merge(
                merged_range.min_row - 1,
                merged_range.max_row - 1,
                merged_range.min_col - 1,
                merged_range.max_col - 1,
            )

    wb_out.save(str(xls_path))
    print(f"[INFO] Converted {xlsx_path.name} -> {xls_path.name}")
    return xls_path


GASTOS_SYSTEM_PROMPT = """Tu eres un contador educado y radicado en Republica Dominicana, experto en temas fiscales, te encargas de procesar recibos de pago y facturas de proveedores para luego ingresarlos en el sistema de contabilidad.

    Tu tarea es extraer la siguiente informacion del recibo/factura y retornarla en formato JSON para ser utilizado en el sistema de contabilidad:

    - nombre: El nombre o razon social del suplidor/proveedor que emite la factura (quien vende, no quien compra). Normalmente aparece en la parte superior del recibo/factura, junto al logo o encabezado. Texto abierto, maximo 255 caracteres.

    - documento: El RNC/cedula(numero de identifcacion de la persona)/numero de pasaporte del suplidor. SOLO digitos, sin guiones ni caracteres especiales (ej: "101702176", "00200078964", "987356102"). Si aparece como "101-70217-6", devolver "101702176".

    - ncf: El NCF (Numero de Comprobante Fiscal). Es un codigo alfanumerico que empieza con una letra (B, E, etc.) seguido de digitos. Ejemplo: E310001987518, B0100014525. Si el valor viene con ceros a la izquierda ANTES de B01 o B02 (ej: "0000000B0100222157"), quita esos ceros y devolver "B0100222157". El B0# puede llegar hasta B09 NO quites ceros internos ni ceros de series E31 u otras (ej: "E310000029838" se deja tal cual).

    **PROTOCOLO OBLIGATORIO PARA LEER EL NCF — sigue estos pasos en orden:**
    Paso 1. Localiza el NCF en la imagen (normalmente aparece junto a la etiqueta "NCF", "Comprobante Fiscal" o "No. Comprobante").
    Paso 2. Lee el codigo CARACTER POR CARACTER de izquierda a derecha, sin saltarte ninguno. Presta atencion especial a zonas con ceros consecutivos: los ceros se parecen entre si y es facil omitir uno o insertar uno de mas.
    Paso 3. Cuenta el total de caracteres del resultado:
       - Serie E31 → DEBE TENER EXACTAMENTE 13 caracteres (1 letra + 12 digitos).
       - Serie B0x → DEBE TENER EXACTAMENTE 11 caracteres (3 letras/digitos + 8 digitos).
    Paso 4. Si el conteo NO coincide con el esperado, o si tuviste cualquier duda al leer un caracter, agrega "ncf" a "campos_dudosos" Y ajusta el score.
    Paso 5. NUNCA insertes ni elimines ceros internos para "cuadrar" el largo; reporta lo que ves aunque el largo quede mal — eso sera detectado automaticamente.
    Ejemplo de error comun: factura tiene "E310000638833" (13 chars) pero el modelo lee "E310000063883" (tambien 13 chars, con un 0 extra insertado y un 3 faltante) — este tipo de transposicion DEBE llevar "ncf" en campos_dudosos.

    - ncf_afectado: NCF modificado/afectado cuando el comprobante es nota de credito (B03) o nota de debito (B04). Maximo 11 caracteres. Si el NCF NO es B03/B04, dejar "".

    - tipo_de_suplidor: Debe ser EXACTAMENTE uno de estos valores:
    Gasto Formal
    Gasto Informal
    Genérico
    Gasto Menor
    Pagos al exterior
    Norma 07-2007
    DGA
    Decreto 139-98
    Regla: si la factura tiene RNC y NCF formal, usar "Gasto Formal"; si no tiene NCF formal, "Gasto Informal".

    - tipo_de_gasto: Debe ser EXACTAMENTE uno de estos valores:
    01-Gasto de personal
    02-Gastos por trabajos, servicios y suministros
    03-Arrendamientos
    04-Gastos de activo fijo
    05-Gastos de representación
    06-Otras deducciones administrativas
    07-Gastos financieros
    08-Gastos extraordinarios
    09-Compras y gastos que forman gastos de la venta
    10-Adquisicion de activos
    11-Gastos de seguros

    - descripcion: Descripcion o comentario del tipo de operacion (ej: "compra de enceres para la casa", "pago de hoteles", "Compra de vegetales",  "COMPRA", "GASOLINA", "MATERIALES", etc). Obligatoria. Maximo 200 caracteres.

    - fecha: Fecha de la transaccion como TEXTO en formato DD/MM/AAAA con dia y mes siempre de 2 digitos (ej: "08/06/2026", "01/11/2025"). Nunca uses "8/06/2026".

    - monto_en_servicios: Monto de la operacion correspondiente a servicios (float). Si la factura es solo bienes, usar 0. Obligatorio.

    - monto_en_bienes: Monto de la operacion correspondiente a bienes / subtotal antes de impuestos si hay ITBIS/SELECTIVO. Si no hay impuestos, usar el total de bienes. Si la factura es solo servicios, usar 0. Obligatorio.

    - itbis: El monto del ITBIS (18%). Buscar cerca del total, etiquetado como "ITBIS", "IVA" o "18%". Valor numerico sin simbolo. Si no aparece, dejar en 0.

    - selectivo: Impuesto selectivo o % LEY si aplica. Normalmente para combustibles y bebidas. Si no aparece, dejar en 0.

    - descuento: Monto del descuento aplicado a la factura, si aparece explicitamente (ej: "Descuento", "Desc.", "Rebaja"). Valor numerico sin simbolo. Si no aparece, dejar en 0.

    - propina: Monto de la propina o servicio, si aparece explicitamente (ej: "Propina", "Servicio", "10% Servicio", "Tip"). Valor numerico sin simbolo. Si no aparece, dejar en 0.

    - moneda: Debe ser EXACTAMENTE uno de estos valores (respeta las mayusculas/minusculas tal cual):
    Peso dominicano
    Dólar americano
    Euros
    Si no se especifica, asume "Peso dominicano".

    - metodo_de_pago: Identificar como uno de:
    + EFECTIVO
    + CHEQUES/TRANSFERENCIAS/DEPÓSITO
    + TARJETA CRÉDITO/DÉBITO
    + COMPRA A CREDITO
    + PERMUTA
    + NOTA DE CREDITO
    + MIXTO

    - campos_dudosos: Array de strings con los NOMBRES EXACTOS (usa las claves JSON, ej. "documento", "ncf", "ncf_afectado", "monto_en_bienes", "monto_en_servicios", "itbis", "fecha", "nombre") de CUALQUIER campo cuyo valor no puedas leer con 100% de certeza en la imagen. Si todos los campos son perfectamente legibles, devuelve un array vacio [].

    - razones_campos_dudosos: Objeto (mapa) donde cada clave es un campo listado en "campos_dudosos" y el valor es un string corto (1-2 oraciones, en espanol) que explica QUE paso y POR QUE ese campo es dudoso. Ejemplo: {"ncf": "El 7mo digito parece 5 o 6 por borrosidad en la zona del NCF.", "documento": "El RNC esta parcialmente cortado en el borde de la foto."}. Si "campos_dudosos" esta vacio, devuelve {}. OBLIGATORIO: todo campo en "campos_dudosos" DEBE tener una entrada aqui con una razon concreta (no generica). Describe el problema visible (borrosidad, glare, digito ambiguo, corte, mancha, longitud inconsistente, etc.).

    **REGLA CRITICA SOBRE CALIDAD DE IMAGEN Y DIGITOS AMBIGUOS (aplica sobre todo a documento, ncf, ncf_afectado, y tambien a monto/itbis/fecha):**
    Los campos numericos/alfanumericos (RNC/cedula, NCF, NCF Afectado, montos) NO tienen contexto de idioma que permita "adivinar" un caracter borroso (a diferencia de una palabra, donde el contexto ayuda a inferir). Por lo tanto, si la foto tiene baja resolucion, esta borrosa, deteriorada, con glare/reflejo, doblada, manchada o cortada justo en la zona de un campo, DEBES tratar cualquier caracter que no se pueda distinguir con 100% de certeza como una duda real, AUNQUE tu "mejor adivinanza" parezca razonable y aunque el resto de la factura sea perfectamente legible.
    Presta especial atencion a pares de digitos que se confunden facilmente cuando la imagen esta degradada: 5 vs 6, 3 vs 8, 0 vs 8 vs 6, 1 vs 7, 2 vs 7, 4 vs 9, 9 vs 8.
    - Reporta siempre tu lectura mas probable en el campo correspondiente (nombre, documento, ncf, etc.), PERO si tienes cualquier duda real sobre uno o mas de sus caracteres por causa de la calidad/deterioro de la imagen, DEBES incluir el nombre de ese campo en "campos_dudosos" Y una razon concreta en "razones_campos_dudosos". Ejemplo: si en el NCF no puedes distinguir si un digito es "5" o "6" (ej. "...613256" vs "...613266"), agrega "ncf" a campos_dudosos y {"ncf": "No se distingue si un digito del NCF es 5 o 6 por borrosidad."} en razones_campos_dudosos.
    - No agregues un campo a "campos_dudosos" solo por incertidumbre de negocio (ej. no saber a que categoria de gasto pertenece); es SOLO para cuando la imagen en si mismo impide leer el valor con certeza total.

    - score: Se calcula EN BASE a "campos_dudosos", no lo decidas de forma independiente:
    + 0 campos en campos_dudosos -> score = 3 (muy seguro)
    + 1 campo en campos_dudosos -> score = 2 (algo seguro)
    + 2 o mas campos en campos_dudosos -> score = 1 (poco seguro)

    ANTES DE ARROJAR EL EXCEL DEBEN SER VALIDADOS ESTOS DATOS 100%, EN CASO DE NO TENER LA CERTEZA  el score debe ser un 2. Si mas de 1 de estos elementos no es 100% certero, se debe asignar un score de 1 de fiabilidad.

    1. Si usa RNC, Confirmar RNC (DEBE TENER 9 CARACTERES) – PERSONA JURIDICA
    2. Si es cedula, Confirmar CEDULA (DEBE TENER 11 CARACTERES) – PERSONA FISICA
    3. NCF — CRITICO: valida el largo caracter por caracter antes de devolver.
       - Serie E31: EXACTAMENTE 13 caracteres totales. Ni 12 ni 14.
       - Serie B0x (B01–B09): EXACTAMENTE 11 caracteres totales. Ni 10 ni 12.
       - Si el largo es incorrecto O si tuviste cualquier duda al leer uno de los digitos, DEBES agregar "ncf" a campos_dudosos.
       - Error tipico a evitar: insertar un cero de mas en una secuencia de ceros consecutivos (ej. leer "E310000063883" en lugar de "E310000638833").

    **Revisar bien la foto por estos valores**  
    4. MONTO: puede aparecer en campos como total, sub-total, Total (sin itbis), Neto, Sub-total excento, etc.
    5. ITBIS
    6. FECHA: Solo se toma la fecha de emision de la factura, se debe ignorar casos similares 'Valido hasta','NFC Vence', 'Fecha', 'Fecha limite pago'.

    Retornar la informacion en formato JSON con las siguientes claves: nombre, documento, ncf, ncf_afectado, tipo_de_suplidor, tipo_de_gasto, descripcion, fecha, monto_en_servicios, monto_en_bienes, itbis, selectivo, descuento, propina, moneda, metodo_de_pago, campos_dudosos, razones_campos_dudosos, score.
    """


def _num(value) -> float:
    """Parse a numeric value, treating blanks as 0.0."""
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        try:
            return float(str(value).replace(",", ""))
        except (TypeError, ValueError):
            return 0.0


def _normalize_documento(value) -> str:
    """Documento must be digits only (no dashes/spaces). Preserves leading zeros."""
    if value is None:
        return ""
    return re.sub(r"\D", "", str(value))


def _normalize_fecha(value) -> str:
    """Normalize date to DD/MM/AAAA text. Accepts common separators and ISO dates."""
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")

    # Already DD/MM/YYYY or D/M/YYYY
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})$", text)
    if m:
        day, month, year = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if year < 100:
            year += 2000
        try:
            return date(year, month, day).strftime("%d/%m/%Y")
        except ValueError:
            return text

    # ISO YYYY-MM-DD
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", text)
    if m:
        year, month, day = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return date(year, month, day).strftime("%d/%m/%Y")
        except ValueError:
            return text

    return text


def _normalize_catalog_value(value: str, options: list[str]) -> str:
    """Return the canonical catalog option, matching case-insensitively / by prefix code."""
    text = (value or "").strip()
    if not text:
        return ""

    lowered = text.lower()
    for option in options:
        if option.lower() == lowered:
            return option

    # Allow "02" or "02-..." style matches for tipo de gasto.
    code_match = re.match(r"^(\d{2})", text)
    if code_match:
        code = code_match.group(1)
        for option in options:
            if option.startswith(code + "-"):
                return option

    # Soft contains match (e.g. "gasto formal" inside a longer phrase).
    for option in options:
        if option.lower() in lowered or lowered in option.lower():
            return option

    return text


def _normalize_moneda(value) -> str:
    """
    Return the canonical "Moneda" value: "Peso dominicano", "Dólar americano"
    or "Euros" - capitalized (first letter only, casing otherwise preserved),
    never all-caps ISO codes like "DOP"/"USD"/"EUR". Defaults to "Peso
    dominicano" when unspecified, per the extraction prompt's own default.
    """
    text = (str(value or "")).strip()
    if not text:
        return "Peso dominicano"

    lowered = text.lower()
    if lowered in MONEDA_ALIASES:
        return MONEDA_ALIASES[lowered]

    for option in MONEDA_OPTIONS:
        if option.lower() == lowered:
            return option

    # Word-boundary substring match (e.g. "USD 500" or "en dolares"), guarded
    # against alnum characters on either side so short aliases can't match
    # inside unrelated words (e.g. "rd$" must not match inside "absurdo").
    for alias, canonical in MONEDA_ALIASES.items():
        pattern = r"(?<![a-z0-9áéíóúñ])" + re.escape(alias) + r"(?![a-z0-9áéíóúñ])"
        if re.search(pattern, lowered):
            return canonical

    return text[:1].upper() + text[1:]


def _normalize_ncf(value) -> str:
    """
    Normalize an NCF: strip whitespace and uppercase only.

    Removing typed nomenclatures (B01, E31, …) on export is owned by the
    frontend ("Remover nomenclaturas NCF") so the user's toggle / series /
    target columns are respected when writing Excel.
    """
    return re.sub(r"\s+", "", str(value or "")).upper()


# Expected exact lengths by NCF series prefix (DGII standard).
# E31 → 13 chars total; B0x (B01–B09) → 11 chars total.
_NCF_EXPECTED_LENGTHS: dict[str, int] = {
    "E31": 13,
    **{f"B0{i}": 11 for i in range(1, 10)},
}


def _ncf_expected_length(ncf: str) -> Optional[int]:
    """Return the canonical character length for a known NCF series, or None."""
    for prefix, length in _NCF_EXPECTED_LENGTHS.items():
        if ncf.startswith(prefix):
            return length
    return None


def _normalize_ncf_afectado(value, ncf: str = "") -> str:
    """Optional; max 11 chars. Required by template when NCF is B03/B04."""
    return _normalize_ncf(value)[:11]


def _to_int_or_none(value) -> Optional[int]:
    """Best-effort int cast; blank/invalid values become None (not 0)."""
    if value in (None, ""):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _parse_catalog_param(raw: Optional[str], param_name: str) -> Optional[list]:
    """Parse a JSON-encoded catalog form field, tolerating blank/invalid input."""
    if not raw:
        return None
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError as e:
        print(f"[WARNING] Could not parse '{param_name}' as JSON: {e}")
        return None
    if not isinstance(parsed, list):
        print(f"[WARNING] '{param_name}' expected a JSON array, got {type(parsed).__name__}")
        return None
    return parsed


def _parse_tax_column_mapping_param(raw: Optional[str]) -> Optional[dict]:
    """Parse the JSON-encoded tax_column_mapping form/query field."""
    if not raw:
        return None
    try:
        parsed = json.loads(raw)
    except json.JSONDecodeError as e:
        print(f"[WARNING] Could not parse 'tax_column_mapping' as JSON: {e}")
        return None
    if not isinstance(parsed, dict):
        print(f"[WARNING] 'tax_column_mapping' expected a JSON object, got {type(parsed).__name__}")
        return None
    return parsed


def _clean_catalog(catalog: Optional[list]) -> list[dict]:
    """
    Normalize a per-client Concepto/Tipo de Pago catalog (from client_documents
    -> document_attributes) to a list of {document_type, document_id, description}.
    Entries without a usable document_id are dropped since they can't be
    written to the export.
    """
    if not catalog:
        return []
    cleaned = []
    for entry in catalog:
        if not isinstance(entry, dict):
            continue
        label = str(entry.get("document_type", "") or "").strip()
        doc_id = _to_int_or_none(entry.get("document_id"))
        if not label or doc_id is None:
            continue
        cleaned.append({
            "document_type": label,
            "document_id": doc_id,
            "description": str(entry.get("description", "") or "").strip(),
        })
    return cleaned


def _clean_business_rules(rules: Optional[list]) -> list[dict]:
    """
    Normalize a per-client business-rules list (from client_business_rules ->
    business_rule_attributes) to {rule_type, rule_value, description}. Unlike
    the Concepto/Tipo de Pago catalogs, these are free-form context hints for
    the AI (not classification options with an ERP id), so entries only need
    a non-empty rule_type to be kept.
    """
    if not rules:
        return []
    cleaned = []
    for entry in rules:
        if not isinstance(entry, dict):
            continue
        rule_type = str(entry.get("rule_type", "") or "").strip()
        if not rule_type:
            continue
        cleaned.append({
            "rule_type": rule_type,
            "rule_value": str(entry.get("rule_value", "") or "").strip(),
            "description": str(entry.get("description", "") or "").strip(),
        })
    return cleaned


def _clean_document_context(entries: Optional[list]) -> list[dict]:
    """
    Normalize a per-client document (client_documents -> document_attributes)
    that's being used purely as CONTEXT for an already-fixed field (e.g.
    tipo_de_gasto), rather than as a list of valid output values. Unlike
    _clean_catalog, entries are kept even without a usable document_id since
    there's no ERP id to write back for this use case - only a non-empty
    document_type (label) is required.
    """
    if not entries:
        return []
    cleaned = []
    for entry in entries:
        if not isinstance(entry, dict):
            continue
        label = str(entry.get("document_type", "") or "").strip()
        if not label:
            continue
        cleaned.append({
            "document_type": label,
            "document_id": _to_int_or_none(entry.get("document_id")),
            "description": str(entry.get("description", "") or "").strip(),
        })
    return cleaned


def _match_catalog_label(label: str, catalog: list[dict]) -> Tuple[str, Optional[int]]:
    """
    Match free text (usually returned by the LLM) against a dynamic per-client
    catalog of {document_type, document_id}. Returns (matched_label, document_id),
    falling back to (original_text, None) when nothing matches.
    """
    text = (label or "").strip()
    if not text or not catalog:
        return text, None

    lowered = text.lower()
    for entry in catalog:
        if entry["document_type"].lower() == lowered:
            return entry["document_type"], entry["document_id"]

    for entry in catalog:
        entry_lower = entry["document_type"].lower()
        if entry_lower in lowered or lowered in entry_lower:
            return entry["document_type"], entry["document_id"]

    return text, None


def _build_gastos_catalog_prompt_block(
    concepto_catalog: list[dict],
    tipo_de_pago_catalog: list[dict],
    concepto_document_comment: str = "",
    tipo_de_pago_document_comment: str = "",
) -> str:
    """
    Build the dynamic, per-client portion of the extraction prompt describing
    the Concepto / Tipo de Pago options available for this specific client.
    Returns "" when neither catalog has usable entries.

    concepto_document_comment / tipo_de_pago_document_comment: optional
    document-level notes (set on the client_documents "Gastos"/"Tipo de
    Pago" container itself, e.g. "comment" column) that apply to ALL options
    in that catalog, on top of each option's own description/comment. Pass
    "" when the document has no comment - the prompt then behaves exactly
    as if this feature didn't exist.
    """
    def _format_options(catalog: list[dict]) -> str:
        return "\n".join(
            f"    - {c['document_type']}" + (f" ({c['description']})" if c["description"] else "")
            for c in catalog
        )

    blocks = []
    if concepto_catalog:
        comment_line = (
            f"    Contexto general de este documento (definido por el usuario, "
            f"aplica a todas las opciones de Concepto de este cliente): "
            f"{concepto_document_comment}\n"
            if concepto_document_comment else ""
        )
        blocks.append(
            "- concepto: Clasifica el gasto segun el CONCEPTO contable especifico "
            "de este cliente. Debe ser EXACTAMENTE uno de estos valores (copia el "
            "texto tal cual, sin agregar nada):\n"
            f"{comment_line}"
            f"{_format_options(concepto_catalog)}\n"
            "    Si ninguno aplica claramente, deja el valor como cadena vacia \"\"."
        )
    if tipo_de_pago_catalog:
        comment_line = (
            f"    Contexto general de este documento (definido por el usuario, "
            f"aplica a todas las opciones de Tipo de Pago de este cliente): "
            f"{tipo_de_pago_document_comment}\n"
            if tipo_de_pago_document_comment else ""
        )
        blocks.append(
            "- tipo_de_pago_erp: Clasifica la forma de pago/registro contable de "
            "este gasto para este cliente especifico. Debe ser EXACTAMENTE uno de "
            "estos valores (copia el texto tal cual, sin agregar nada):\n"
            f"{comment_line}"
            f"{_format_options(tipo_de_pago_catalog)}\n"
            "    Este campo es OBLIGATORIO y NUNCA debe quedar vacio: si ninguno "
            "coincide perfectamente, elige el valor de la lista que mas se "
            "aproxime a la forma de pago detectada en el documento."
        )

    if not blocks:
        return ""

    return (
        "\n\nADICIONAL - Catalogos especificos de este cliente (dinamicos, "
        "varian por cliente, NO uses conocimiento general para esto):\n"
        + "\n".join(blocks)
        + "\n\nIncluye 'concepto' y 'tipo_de_pago_erp' como claves adicionales "
        "en el JSON de salida. 'concepto' puede ser cadena vacia \"\" si "
        "ninguna opcion aplica, pero 'tipo_de_pago_erp' SIEMPRE debe llevar "
        "uno de los valores listados (nunca cadena vacia)."
    )


def _build_gastos_business_rules_prompt_block(business_rules: list[dict]) -> str:
    """
    Build the optional, per-client "business rules" portion of the
    extraction prompt: free-form context (exceptions, conventions,
    classification hints, etc.) that helps the AI make better decisions for
    this specific client. Unlike the Concepto/Tipo de Pago catalogs, these
    are NOT a fixed set of valid output values - just contextual guidance.
    Returns "" when there are no usable rules.
    """
    if not business_rules:
        return ""

    lines = []
    for rule in business_rules:
        label = rule["rule_type"]
        if rule.get("rule_value"):
            label += f" ({rule['rule_value']})"
        if rule.get("description"):
            lines.append(f"    - {label}: {rule['description']}")
        else:
            lines.append(f"    - {label}")

    return (
        "\n\nADICIONAL - Reglas de negocio (organizacion y/o cliente; "
        "contexto para ayudarte a tomar mejores decisiones al clasificar "
        "y extraer este documento; NO son valores fijos que debas copiar "
        "literalmente, solo guian tu criterio):\n"
        + "\n".join(lines)
    )


def _build_gastos_tipo_de_gasto_context_block(
    context: list[dict], document_comment: str = ""
) -> str:
    """
    Build optional, per-client context to help the AI choose among the
    FIXED tipo_de_gasto options (TIPO_DE_GASTO_OPTIONS, baked into
    GASTOS_SYSTEM_PROMPT) - typically a client_documents container the user picked
    specifically to describe how THIS client's suppliers/categories map to
    those 11 fixed options. This block NEVER introduces new tipo_de_gasto
    values; it only guides which of the existing 11 fits best. Returns ""
    when there's nothing to add.
    """
    if not context and not document_comment:
        return ""

    lines = []
    if document_comment:
        lines.append(
            f"    Contexto general de este documento (definido por el "
            f"usuario): {document_comment}"
        )
    for entry in context:
        label = entry["document_type"]
        if entry.get("document_id") is not None:
            label += f" (ref. {entry['document_id']})"
        if entry.get("description"):
            lines.append(f"    - {label}: {entry['description']}")
        else:
            lines.append(f"    - {label}")

    if not lines:
        return ""

    return (
        "\n\nADICIONAL - Contexto especifico de este cliente para elegir "
        "'tipo_de_gasto' (usa esto UNICAMENTE como ayuda para decidir cual "
        "de las 11 opciones fijas de tipo_de_gasto (arriba) aplica mejor; "
        "NUNCA inventes un valor nuevo ni copies este texto como respuesta "
        "Si no estas 100% seguro de la categoria de tipo de gasto, influir en tu decision para el Score de fiabilidad"
        "- el resultado debe seguir siendo EXACTAMENTE uno de los 11 "
        "valores fijos listados):\n" + "\n".join(lines)
    )


# Field names the LLM is allowed to flag in "campos_dudosos". Anything else
# it returns (typos, business-logic fields, etc.) is dropped defensively.
# "concepto_id" is not flagged by the LLM (it resolves the label post-extraction)
# but is injected by _normalize_gastos_extracted when no catalog match is found.
CAMPOS_DUDOSOS_VALID_KEYS = {
    "nombre", "documento", "ncf", "ncf_afectado", "tipo_de_suplidor",
    "tipo_de_gasto", "descripcion", "fecha", "monto_en_servicios",
    "monto_en_bienes", "itbis", "selectivo", "descuento", "propina",
    "moneda", "metodo_de_pago", "concepto_id",
}

# Fallback hover text when a field is flagged but the model omitted a reason.
DEFAULT_DUDOSO_REASON = (
    "Imagen poco clara: no se pudo leer este valor con certeza total."
)

# Reasons for server-side flags (not produced by the LLM).
SERVER_DUDOSO_REASONS = {
    "ncf": (
        "El NCF no tiene la longitud esperada para su serie "
        "(E31 = 13 caracteres, B0x = 11); validar dígito por dígito."
    ),
    "concepto_id": (
        "No se encontró coincidencia en el catálogo de conceptos; "
        "seleccionar manualmente."
    ),
}


def _campo_dudoso_key(item) -> str:
    """Extract a field key from a string or {campo/field/nombre: ...} object."""
    if isinstance(item, dict):
        raw = (
            item.get("campo")
            or item.get("field")
            or item.get("nombre")
            or item.get("key")
            or ""
        )
        return str(raw or "").strip().lower()
    return str(item or "").strip().lower()


def _normalize_campos_dudosos(value) -> list[str]:
    """
    Coerce the LLM's 'campos_dudosos' into a de-duplicated list of known field
    keys. Accepts plain strings or objects like {"campo": "ncf", "razon": "..."}.
    """
    if not value:
        return []
    if isinstance(value, str):
        value = [value]
    if not isinstance(value, (list, tuple, set)):
        return []
    seen: list[str] = []
    for item in value:
        key = _campo_dudoso_key(item)
        if key in CAMPOS_DUDOSOS_VALID_KEYS and key not in seen:
            seen.append(key)
    return seen


def _normalize_razones_campos_dudosos(
    value,
    campos_dudosos: list[str],
    *,
    campos_dudosos_raw=None,
) -> dict[str, str]:
    """
    Build {field: reason} for every flagged field.

    Accepts:
    - dict mapping field -> reason string
    - list of {campo, razon} objects (also pulled from campos_dudosos_raw)
    Ensures every key in campos_dudosos has a non-empty reason (fallback if missing).
    """
    reasons: dict[str, str] = {}

    def _put(field: str, reason) -> None:
        key = str(field or "").strip().lower()
        text = str(reason or "").strip()
        if key in CAMPOS_DUDOSOS_VALID_KEYS and text and key not in reasons:
            reasons[key] = text

    if isinstance(value, dict):
        for field, reason in value.items():
            _put(field, reason)
    elif isinstance(value, (list, tuple)):
        for item in value:
            if isinstance(item, dict):
                _put(
                    _campo_dudoso_key(item),
                    item.get("razon")
                    or item.get("reason")
                    or item.get("motivo")
                    or item.get("mensaje"),
                )

    # Also harvest reasons embedded in campos_dudosos as objects.
    if isinstance(campos_dudosos_raw, (list, tuple)):
        for item in campos_dudosos_raw:
            if isinstance(item, dict):
                _put(
                    _campo_dudoso_key(item),
                    item.get("razon")
                    or item.get("reason")
                    or item.get("motivo")
                    or item.get("mensaje"),
                )

    for field in campos_dudosos:
        if field not in reasons:
            reasons[field] = DEFAULT_DUDOSO_REASON

    # Drop reasons for fields that are not actually flagged.
    return {k: reasons[k] for k in campos_dudosos if k in reasons}


def _score_from_campos_dudosos(campos_dudosos: list[str]) -> int:
    """
    Deterministic score derived from how many fields the LLM flagged as
    uncertain due to image quality: 0 -> 3, 1 -> 2, 2+ -> 1. Mirrors the
    scoring rule described in GASTOS_SYSTEM_PROMPT, but enforced in code so the
    model can't self-report a confident score (e.g. 3) while separately
    admitting a field is ambiguous.
    """
    count = len(campos_dudosos)
    if count == 0:
        return 3
    if count == 1:
        return 2
    return 1


def prepare_gastos_export_row(data: dict) -> dict:
    """
    Normalize a receipt dict to Carga Masiva template rules before writing Excel.
    Safe to call on both freshly extracted and user-edited payloads.
    """
    ncf = _normalize_ncf(data.get("ncf", ""))
    nombre = str(data.get("nombre", "") or "").strip()[:255]
    descripcion = str(data.get("descripcion", "") or "").strip()[:200]

    # Server-side NCF length validation.
    # Any mismatch (too long OR too short) means the AI likely mis-read a digit
    # or inserted/dropped a zero in a run of consecutive zeros.
    # - Over-length: truncate to cap so the template column never exceeds the
    #   DGII limit (even though the value may still be wrong after truncation).
    # - Under-length: leave as-is; we cannot know which digit is missing.
    # Either way: flag "ncf" in campos_dudosos so the score drops and the user
    # is prompted to review the field manually.
    server_dubious: list[str] = []
    expected_ncf_len = _ncf_expected_length(ncf)
    if expected_ncf_len and len(ncf) != expected_ncf_len:
        if len(ncf) > expected_ncf_len:
            print(
                f"[WARNING] NCF '{ncf}' is {len(ncf)} chars (expected "
                f"{expected_ncf_len}); truncating to cap and flagging as dubious."
            )
            ncf = ncf[:expected_ncf_len]
        else:
            print(
                f"[WARNING] NCF '{ncf}' is only {len(ncf)} chars (expected "
                f"{expected_ncf_len}); flagging as dubious."
            )
        server_dubious.append("ncf")

    ncf_afectado = _normalize_ncf_afectado(data.get("ncf_afectado", ""), ncf)

    if (
        any(ncf.startswith(prefix) for prefix in NCF_AFECTADO_REQUIRED_PREFIXES)
        and not ncf_afectado
    ):
        print(
            f"[WARNING] NCF {ncf} requires NCF Afectado (B03/B04) but none was provided"
        )

    campos_dudosos_raw = data.get("campos_dudosos")
    campos_dudosos = _normalize_campos_dudosos(campos_dudosos_raw)
    # Merge server-side findings (e.g. over-length NCF) that the model missed.
    for field in server_dubious:
        if field not in campos_dudosos:
            campos_dudosos.append(field)

    razones_campos_dudosos = _normalize_razones_campos_dudosos(
        data.get("razones_campos_dudosos"),
        campos_dudosos,
        campos_dudosos_raw=campos_dudosos_raw,
    )
    # Prefer the server-authored reason when we ourselves flagged the field.
    for field in server_dubious:
        if field in SERVER_DUDOSO_REASONS:
            razones_campos_dudosos[field] = SERVER_DUDOSO_REASONS[field]
    # concepto_id is injected post-extraction (catalog miss); if still on the
    # generic fallback, upgrade to the catalog-specific message.
    if (
        "concepto_id" in razones_campos_dudosos
        and razones_campos_dudosos["concepto_id"] == DEFAULT_DUDOSO_REASON
    ):
        razones_campos_dudosos["concepto_id"] = SERVER_DUDOSO_REASONS["concepto_id"]

    reported_score = _to_int_or_none(data.get("score")) or 0
    if reported_score > 0:
        # Cap (never raise) the model's self-reported score using the
        # deterministic rule, so a model that flags an ambiguous field but
        # still reports score=3 gets overridden down to what the rule says.
        score = min(reported_score, _score_from_campos_dudosos(campos_dudosos))
    else:
        # 0 is the failure/empty-extraction sentinel elsewhere - leave untouched.
        score = reported_score

    return {
        "nombre": nombre,
        "documento": _normalize_documento(data.get("documento", "")),
        "ncf": ncf,
        "ncf_afectado": ncf_afectado,
        "tipo_de_suplidor": _normalize_catalog_value(
            str(data.get("tipo_de_suplidor", "") or ""), TIPO_DE_SUPLIDOR_OPTIONS
        ),
        "tipo_de_gasto": _normalize_catalog_value(
            str(data.get("tipo_de_gasto", "") or ""), TIPO_DE_GASTO_OPTIONS
        ),
        "descripcion": descripcion,
        "fecha": _normalize_fecha(data.get("fecha", "")),
        "monto_en_servicios": _num(data.get("monto_en_servicios")),
        "monto_en_bienes": _num(data.get("monto_en_bienes")),
        "itbis": _num(data.get("itbis")),
        "selectivo": _num(data.get("selectivo")),
        "descuento": _num(data.get("descuento")),
        "propina": _num(data.get("propina")),
        "moneda": _normalize_moneda(data.get("moneda", "")),
        "metodo_de_pago": str(data.get("metodo_de_pago", "") or "").strip(),
        "concepto_id": _to_int_or_none(data.get("concepto_id")),
        "tipo_de_pago_id": _to_int_or_none(data.get("tipo_de_pago_id")),
        "filename": data.get("filename", "") or "",
        "score": score,
        "campos_dudosos": campos_dudosos,
        "razones_campos_dudosos": razones_campos_dudosos,
    }


def _normalize_gastos_for_key(value) -> str:
    """Collapse whitespace and lowercase, for building gastos duplicate-detection keys."""
    return re.sub(r"\s+", " ", str(value or "").strip().lower())


def _gastos_receipt_identity_key(data: dict) -> Optional[str]:
    """
    Build a signature identifying the underlying receipt/invoice represented
    by an extracted-data dict, so the SAME receipt scanned/uploaded more than
    once (as a different file, photo angle, or duplicate page) can be
    detected even when the file bytes differ.

    - When both `documento` (supplier RNC) and `ncf` are present, the pair is
      used: an NCF is unique per supplier under DR tax law, so this is the
      strongest possible duplicate signal.
    - Otherwise (e.g. informal/no-NCF receipts) falls back to
      documento/nombre + fecha + monto totals - a weaker but still
      reasonable signal for receipts without a fiscal number.
    - Returns None when there isn't enough extracted signal to safely judge
      duplication (e.g. an almost-empty/failed extraction) - under-detecting
      is preferable to wrongly collapsing two unrelated blank results.
    """
    prepared = prepare_gastos_export_row(data)
    documento = _normalize_gastos_for_key(prepared.get("documento"))
    ncf = _normalize_gastos_for_key(prepared.get("ncf"))
    nombre = _normalize_gastos_for_key(prepared.get("nombre"))
    fecha = _normalize_gastos_for_key(prepared.get("fecha"))
    supplier = documento or nombre

    if ncf and supplier:
        return f"ncf:{supplier}|{ncf}"

    if not supplier or not fecha:
        return None

    monto_servicios = round(_num(prepared.get("monto_en_servicios")), 2)
    monto_bienes = round(_num(prepared.get("monto_en_bienes")), 2)
    if monto_servicios == 0 and monto_bienes == 0:
        return None

    return f"amt:{supplier}|{fecha}|{monto_servicios}|{monto_bienes}"


# Template header aliases (headers are matched lowercased + stripped).
# Note: the official template misspells "Decripcion" and trailing-spaces some headers.
GASTOS_EXCEL_FIELD_MAPPINGS = {
    "nombre": ["nombre", "nombre del proveedor", "nombre del suplidor", "proveedor", "suplidor"],
    "documento": ["documento", "rnc", "nif", "ruc"],
    "ncf": ["ncf", "comprobante fiscal"],
    "ncf_afectado": ["ncf afectado", "nfc afectado"],
    "tipo_de_suplidor": ["tipo de suplidor", "tipo suplidor"],
    "tipo_de_gasto": ["tipo de gasto", "tipo gasto"],
    "descripcion": ["decripcion", "descripcion", "concepto", "detalle"],
    "fecha": ["fecha", "date", "fecha factura"],
    "monto_en_servicios": ["monto en servicios", "monto servicios"],
    "monto_en_bienes": ["monto en bienes", "monto bienes"],
    # The template only exposes 5 generic tax slots (Impuesto 1..5). WHICH
    # semantic amount (itbis/selectivo/descuento/propina) lands in which slot
    # is decided per-client by GASTOS_TAX_COLUMN_FIELDS / resolve_gastos_tax_columns below,
    # so these keys (not "itbis"/"selectivo" directly) own the header aliases.
    "impuesto_1": ["impuesto 1"],
    "impuesto_2": ["impuesto 2"],
    "impuesto_3": ["impuesto 3"],
    "impuesto_4": ["impuesto 4"],
    "impuesto_5": ["impuesto 5"],
    "moneda": ["moneda", "currency", "divisa"],
    "metodo_de_pago": ["forma de pago", "metodo de pago", "forma pago"],
    "concepto_id": ["concepto id"],
    "tipo_de_pago_id": ["tipo de pago id", "tipo pago id"],
}

GASTOS_EXCEL_TEXT_FIELDS = [
    "nombre", "documento", "ncf", "ncf_afectado", "tipo_de_suplidor", "tipo_de_gasto",
    "descripcion", "fecha", "moneda", "metodo_de_pago",
]
GASTOS_EXCEL_NUMERIC_FIELDS = [
    "monto_en_servicios", "monto_en_bienes",
    "impuesto_1", "impuesto_2", "impuesto_3", "impuesto_4", "impuesto_5",
]
GASTOS_EXCEL_INT_FIELDS = ["concepto_id", "tipo_de_pago_id"]

    # Amounts that can be routed into one of the 5 "Impuesto" export columns.
    # Matches composables/useClientTaxColumnMapping.ts's GASTOS_TAX_COLUMN_FIELDS.
GASTOS_TAX_COLUMN_FIELDS = ["itbis", "selectivo", "descuento", "propina"]
# Fallback used when the client has no configured mapping yet, preserving
# the export's original (pre-Descuento/Propina) behavior.
GASTOS_DEFAULT_TAX_COLUMN_MAPPING = {"itbis": 1, "selectivo": 2, "descuento": None, "propina": None}


def _clean_gastos_tax_column_mapping(mapping: Optional[dict]) -> dict:
    """
    Normalize a client's {field: impuesto_slot} mapping. Unknown fields are
    dropped, out-of-range/duplicate slots fall back to "unmapped" (None) for
    that field so a bad client config never silently overwrites another
    field's column, and any field missing from `mapping` keeps its default.
    """
    resolved = dict(GASTOS_DEFAULT_TAX_COLUMN_MAPPING)
    if isinstance(mapping, dict):
        seen_slots: set[int] = set()
        for field in GASTOS_TAX_COLUMN_FIELDS:
            if field not in mapping:
                continue
            slot = _to_int_or_none(mapping.get(field))
            if slot is not None and 1 <= slot <= 5 and slot not in seen_slots:
                resolved[field] = slot
                seen_slots.add(slot)
            else:
                resolved[field] = None
    return resolved


def resolve_gastos_tax_columns(row: dict, tax_column_mapping: Optional[dict]) -> dict:
    """
    Given a prepare_gastos_export_row()-shaped dict (with itbis/selectivo/descuento/
    propina already normalized to floats), return {impuesto_1..impuesto_5:
    float} placing each amount in the client-configured slot. Slots with no
    field mapped to them default to 0.0.
    """
    mapping = _clean_gastos_tax_column_mapping(tax_column_mapping)
    slots = {f"impuesto_{n}": 0.0 for n in range(1, 6)}
    for field in GASTOS_TAX_COLUMN_FIELDS:
        slot = mapping.get(field)
        if slot:
            slots[f"impuesto_{slot}"] = _num(row.get(field))
    return slots
# Fecha is exported as plain text (DD/MM/YYYY), not an Excel date serial.
GASTOS_EXCEL_DATE_FIELDS: list[str] = []

def _normalize_gastos_extracted(
    extracted: dict,
    filename: str,
    concepto_catalog: Optional[list[dict]] = None,
    tipo_de_pago_catalog: Optional[list[dict]] = None,
) -> dict:
    """
    Normalize a single extracted-data dict to our expected shape. When
    per-client catalogs are provided, resolves the LLM's free-text
    'concepto' / 'tipo_de_pago_erp' picks into their ERP integer ids
    (concepto_id / tipo_de_pago_id) via _match_catalog_label.
    """
    _, concepto_id = _match_catalog_label(
        str(extracted.get("concepto", "") or ""), concepto_catalog or []
    )

    # "Tipo de Pago Id" is a required field for the destination CRM, so it
    # must NEVER be left blank once the client has a Tipo de Pago catalog
    # selected. Try the LLM's dedicated guess first, then fall back to
    # matching the raw "metodo_de_pago" text, and finally default to the
    # first catalog entry so we always emit a valid id.
    _, tipo_de_pago_id = _match_catalog_label(
        str(extracted.get("tipo_de_pago_erp", "") or ""), tipo_de_pago_catalog or []
    )
    if tipo_de_pago_id is None and tipo_de_pago_catalog:
        _, tipo_de_pago_id = _match_catalog_label(
            str(extracted.get("metodo_de_pago", "") or ""), tipo_de_pago_catalog
        )
    if tipo_de_pago_id is None and tipo_de_pago_catalog:
        tipo_de_pago_id = tipo_de_pago_catalog[0]["document_id"]

    # Concepto is mandatory when the client has a catalog configured.
    # If no match was found, flag it so the score drops and the cell is
    # highlighted for manual review in the frontend.
    extra_dudosos = list(extracted.get("campos_dudosos") or [])
    # Preserve whatever shape the model returned; prepare_gastos_export_row
    # normalizes dict / list-of-objects and upgrades concepto_id's fallback
    # reason to the catalog-specific message.
    extra_razones = extracted.get("razones_campos_dudosos")
    if concepto_catalog and concepto_id is None and "concepto_id" not in {
        _campo_dudoso_key(x) for x in extra_dudosos
    }:
        print(
            f"[WARNING] '{filename}': concepto catalog present but no match found "
            f"for '{extracted.get('concepto', '')}'; flagging concepto_id as dubious."
        )
        extra_dudosos.append("concepto_id")
        if isinstance(extra_razones, dict):
            extra_razones = {
                **extra_razones,
                "concepto_id": SERVER_DUDOSO_REASONS["concepto_id"],
            }

    prepared = prepare_gastos_export_row({
        **extracted,
        "campos_dudosos": extra_dudosos,
        "razones_campos_dudosos": extra_razones,
        "concepto_id": concepto_id,
        "tipo_de_pago_id": tipo_de_pago_id,
        "filename": filename,
        "score": extracted.get("score", 0) or 0,
    })
    return prepared


def _empty_gastos_extracted(filename: str, descripcion: str = "") -> dict:
    """Return an empty/default extracted-data dict for failed processing."""
    return prepare_gastos_export_row({
        "descripcion": descripcion,
        "filename": filename,
        "score": 0,
    })


# _is_retryable_error, PDF_MAX_PAGES, PDF_RENDER_DPI, render_pdf_to_images,
# load_file_as_images, _parse_json_loose, and gemini_semaphore are
# imported from shared_utils at the top of this file.


def process_gastos_with_gemini(
    file_content: bytes,
    filename: str,
    max_retries: int = 3,
    concepto_catalog: Optional[list[dict]] = None,
    tipo_de_pago_catalog: Optional[list[dict]] = None,
    concepto_document_comment: str = "",
    tipo_de_pago_document_comment: str = "",
    business_rules: Optional[list[dict]] = None,
    tipo_de_gasto_context: Optional[list[dict]] = None,
    tipo_de_gasto_document_comment: str = "",
    model_name: Optional[str] = None,
    spend_ctx: Optional[dict] = None,
) -> Tuple[dict, Optional[dict]]:
    """
    Process a single file (image or PDF) with the selected vision model.
    Extracts receipt/invoice data according to Dominican Republic accounting
    standards. PDFs are rasterized to images first so vision-only models
    (Gemma) can process them. Includes retry logic for rate limiting and
    transient errors.

    Returns (extracted_data, usage_record_or_None).

    model_name: optional client-selected model (thinking level). Falls back
    to INDIVIDUAL_MODEL when omitted or not whitelisted.

    spend_ctx: optional {thinking_level, organization_id, user_id, client_id}
    used to price and persist token usage after a successful Gemini call.

    concepto_catalog / tipo_de_pago_catalog: optional per-client lists of
    {document_type, document_id, description} (already cleaned via
    _clean_catalog) used to classify the receipt into that client's Concepto
    / Tipo de Pago ERP ids dynamically.

    concepto_document_comment / tipo_de_pago_document_comment: optional
    document-level comments for extra context. "" when the document has none.

    business_rules: optional per-client list of {rule_type, rule_value,
    description} (from client_business_rules / business_rule_attributes,
    already cleaned via _clean_business_rules) - free-form context to help
    the AI make better decisions for this client, independent of the
    Concepto/Tipo de Pago catalogs.

    tipo_de_gasto_context / tipo_de_gasto_document_comment: optional
    per-client document (any client_documents container the user picks,
    already cleaned via _clean_document_context) used PURELY as context to
    help choose among the FIXED tipo_de_gasto options - it never introduces
    new tipo_de_gasto values.
    """
    import time

    model_id = resolve_inference_model(model_name, default=INDIVIDUAL_MODEL)

    catalog_block = _build_gastos_catalog_prompt_block(
        concepto_catalog or [], tipo_de_pago_catalog or [],
        concepto_document_comment=concepto_document_comment,
        tipo_de_pago_document_comment=tipo_de_pago_document_comment,
    )
    catalog_block += _build_gastos_business_rules_prompt_block(business_rules or [])
    catalog_block += _build_gastos_tipo_de_gasto_context_block(
        tipo_de_gasto_context or [], document_comment=tipo_de_gasto_document_comment
    )

    print(
        f"[INFO] [GEMINI-SINGLE] Starting processing for '{filename}' "
        f"({len(file_content)} bytes, model={model_id})"
    )

    # Check if API key is configured
    if not GEMINI_API_KEY:
        print(f"[ERROR] [GEMINI-SINGLE] Cannot process {filename}: GEMINI_API_KEY not configured")
        return _empty_gastos_extracted(filename, descripcion="ERROR: API key not configured"), None

    file_extension = Path(filename).suffix.lower() if filename else ""
    if file_extension not in (".png", ".jpg", ".jpeg", ".pdf"):
        print(f"[ERROR] [GEMINI-SINGLE] Unsupported file type: {file_extension}")
        return _empty_gastos_extracted(
            filename, descripcion=f"Unsupported file type: {file_extension}"
        ), None

    page_images = load_file_as_images(file_content, filename)
    print(f"[DEBUG] [GEMINI-SINGLE] '{filename}': loaded {len(page_images)} page image(s)")
    if not page_images:
        print(f"[ERROR] [GEMINI-SINGLE] '{filename}': no page images produced, aborting before calling Gemini")
        return _empty_gastos_extracted(
            filename,
            descripcion=(
                "Could not render PDF pages"
                if file_extension == ".pdf"
                else "Could not open image"
            ),
        ), None

    if file_extension == ".pdf" and len(page_images) > 1:
        prompt = (
            f"{GASTOS_SYSTEM_PROMPT}{catalog_block}\n\n"
            f"NOTA: Te envio {len(page_images)} paginas de un mismo recibo/"
            f"factura en PDF. Trata las paginas como un solo documento y "
            f"retorna SOLO un objeto JSON valido sin texto adicional, sin "
            f"markdown, sin explicaciones."
        )
    else:
        prompt = (
            f"{GASTOS_SYSTEM_PROMPT}{catalog_block}\n\nExtrae la informacion del "
            f"recibo/factura en la imagen y retorna SOLO un objeto JSON "
            f"valido sin texto adicional, sin markdown, sin explicaciones."
        )

    last_error = None

    for attempt in range(max_retries):
        try:
            print(
                f"[INFO] [GEMINI-SINGLE] '{filename}': calling Gemini "
                f"(attempt {attempt + 1}/{max_retries}, model={model_id}, "
                f"images={len(page_images)}, prompt_chars={len(prompt)})"
            )
            content_parts = [prompt, *page_images]
            ctx = spend_ctx or {}
            level = ctx.get("thinking_level") or resolve_thinking_level(
                None, model_id
            )

            # Structured JSON for all levels; moderado also sends thinkingBudget.
            response = generate_inference_content(
                model_id,
                content_parts,
                thinking_level=level,
                max_output_tokens=4096,
                temperature=0.1,
            )

            if not response or not getattr(response, "text", None):
                raise ValueError("Empty response from Gemini API")

            usage_record = record_usage_from_response(
                response,
                model=model_id,
                source="gastos_single",
                thinking_level=level,
                organization_id=ctx.get("organization_id"),
                user_id=ctx.get("user_id"),
                client_id=ctx.get("client_id"),
                metadata={"filename": filename},
            )

            extracted_data = _parse_json_loose(response.text)
            if isinstance(extracted_data, list):
                # Rare: model returns a 1-element array instead of an object.
                extracted_data = next(
                    (item for item in extracted_data if isinstance(item, dict)),
                    {},
                )
            if not isinstance(extracted_data, dict):
                raise ValueError(
                    f"Expected JSON object from single response, got "
                    f"{type(extracted_data).__name__}"
                )

            normalized = _normalize_gastos_extracted(
                extracted_data, filename,
                concepto_catalog=concepto_catalog,
                tipo_de_pago_catalog=tipo_de_pago_catalog,
            )
            is_empty = not any(
                normalized.get(k) for k in
                ("nombre", "documento", "ncf", "tipo_de_suplidor", "fecha")
            )
            print(
                f"[INFO] [GEMINI-SINGLE] '{filename}': success, score="
                f"{normalized.get('score')}, empty_fields={is_empty}"
            )
            if is_empty:
                print(
                    f"[WARNING] [GEMINI-SINGLE] '{filename}': Gemini returned a "
                    f"parseable but essentially empty result - the image content "
                    f"may be blank/unreadable, or the file sent doesn't match "
                    f"what was expected."
                )
            return normalized, usage_record

        except Exception as e:
            last_error = e
            print(
                f"[ERROR] [GEMINI-SINGLE] '{filename}': attempt {attempt + 1}/"
                f"{max_retries} failed: {e}"
            )
            if _is_retryable_error(e) and attempt < max_retries - 1:
                print(f"[INFO] [GEMINI-SINGLE] '{filename}': retryable error, backing off before retry")
                time.sleep(2 ** attempt)
            elif not _is_retryable_error(e):
                break

    print(f"[ERROR] [GEMINI-SINGLE] Gemini processing failed for {filename}: {last_error}")
    return _empty_gastos_extracted(filename), None


def _process_gastos_files_parallel(
    files: List[Tuple[bytes, str]],
    *,
    concepto_catalog: Optional[list[dict]] = None,
    tipo_de_pago_catalog: Optional[list[dict]] = None,
    concepto_document_comment: str = "",
    tipo_de_pago_document_comment: str = "",
    business_rules: Optional[list[dict]] = None,
    tipo_de_gasto_context: Optional[list[dict]] = None,
    tipo_de_gasto_document_comment: str = "",
    model_name: Optional[str] = None,
    spend_ctx: Optional[dict] = None,
) -> Tuple[List[dict], Optional[dict]]:
    """
    Process each file with process_gastos_with_gemini in parallel (capped by
    GEMINI_MAX_CONCURRENT). Preserves input order in the returned results.
    """
    from concurrent.futures import ThreadPoolExecutor, as_completed

    if not files:
        return [], None

    results: List[Optional[dict]] = [None] * len(files)
    usages: List[dict] = []

    def _run(index: int, content: bytes, filename: str):
        return index, process_gastos_with_gemini(
            content,
            filename,
            concepto_catalog=concepto_catalog,
            tipo_de_pago_catalog=tipo_de_pago_catalog,
            concepto_document_comment=concepto_document_comment,
            tipo_de_pago_document_comment=tipo_de_pago_document_comment,
            business_rules=business_rules,
            tipo_de_gasto_context=tipo_de_gasto_context,
            tipo_de_gasto_document_comment=tipo_de_gasto_document_comment,
            model_name=model_name,
            spend_ctx=spend_ctx,
        )

    workers = min(GEMINI_MAX_CONCURRENT, len(files))
    print(
        f"[INFO] [GEMINI-PARALLEL] Processing {len(files)} file(s) with "
        f"{workers} worker(s), model={resolve_inference_model(model_name, default=INDIVIDUAL_MODEL)}"
    )
    with ThreadPoolExecutor(max_workers=workers) as pool:
        futures = [
            pool.submit(_run, i, content, filename)
            for i, (content, filename) in enumerate(files)
        ]
        for fut in as_completed(futures):
            index, (data, usage) = fut.result()
            results[index] = data
            if usage:
                usages.append(usage)

    ordered = [r if r is not None else _empty_gastos_extracted(fn) for r, (_, fn) in zip(results, files)]
    return ordered, merge_usage_records(usages)


def process_gastos_batch_with_gemini(
    files: List[Tuple[bytes, str]],
    max_retries: int = 3,
    concepto_catalog: Optional[list[dict]] = None,
    tipo_de_pago_catalog: Optional[list[dict]] = None,
    concepto_document_comment: str = "",
    tipo_de_pago_document_comment: str = "",
    business_rules: Optional[list[dict]] = None,
    tipo_de_gasto_context: Optional[list[dict]] = None,
    tipo_de_gasto_document_comment: str = "",
    model_name: Optional[str] = None,
    spend_ctx: Optional[dict] = None,
) -> Tuple[List[dict], Optional[dict]]:
    """
    Process MULTIPLE files (images and/or PDFs) with the selected vision model
    in a SINGLE API call.

    Returns (results, usage_record_or_None).

    files: list of (file_content, filename) tuples. Supports png/jpg/jpeg and
    pdf - PDFs are rasterized to images on the fly. Each file becomes ONE entry
    in the result list, even if its PDF spans multiple pages.

    model_name: optional client-selected model (thinking level). Falls back
    to BATCH_MODEL when omitted or not whitelisted.

    spend_ctx: optional {thinking_level, organization_id, user_id, client_id}
    used to price and persist token usage after a successful Gemini call.

    concepto_catalog / tipo_de_pago_catalog: optional per-client lists of
    {document_type, document_id, description} used to classify every
    document in the batch into that client's Concepto / Tipo de Pago ERP ids.

    concepto_document_comment / tipo_de_pago_document_comment: optional
    document-level comments for extra context. "" when the document has none.

    business_rules: optional per-client list of {rule_type, rule_value,
    description} (already cleaned via _clean_business_rules) - free-form
    context to help the AI make better decisions for this client, applied to
    every document in the batch.

    tipo_de_gasto_context / tipo_de_gasto_document_comment: optional
    per-client document (already cleaned via _clean_document_context) used
    PURELY as context to help choose among the FIXED tipo_de_gasto options
    for every document in the batch - it never introduces new values.

    Returns a list of normalized extracted-data dicts in the same order as
    input. Falls back to per-file processing if the batch call cannot be parsed
    back into the expected number of items.

    Gemma is unreliable at multi-doc JSON arrays, so multi-file Gemma requests
    skip the single-shot batch and run parallel per-file calls instead.

    Large multi-doc requests (> MAX_GEMINI_MULTI_DOC) are split into sequential
    sub-batches — packing 15–25 images into one generateContent call routinely
    truncates the JSON array (1 of N results) or times out.
    """
    import time

    if not files:
        return [], None

    model_id = resolve_inference_model(model_name, default=BATCH_MODEL)

    parallel_kwargs = dict(
        concepto_catalog=concepto_catalog,
        tipo_de_pago_catalog=tipo_de_pago_catalog,
        concepto_document_comment=concepto_document_comment,
        tipo_de_pago_document_comment=tipo_de_pago_document_comment,
        business_rules=business_rules,
        tipo_de_gasto_context=tipo_de_gasto_context,
        tipo_de_gasto_document_comment=tipo_de_gasto_document_comment,
        model_name=model_id,
        spend_ctx=spend_ctx,
    )

    # Gemma often emits junk before/around multi-doc JSON arrays (Extra data),
    # which forces a slow sequential fallback. Prefer parallel singles up front.
    if len(files) > 1 and model_id.startswith("gemma"):
        print(
            f"[INFO] [GEMINI-BATCH] model={model_id} is Gemma — skipping multi-doc "
            f"batch API; using parallel per-file for {len(files)} file(s)"
        )
        return _process_gastos_files_parallel(files, **parallel_kwargs)

    # Cap single-shot multi-doc size. Override with MAX_GEMINI_MULTI_DOC.
    max_multi = max(1, int(os.getenv("MAX_GEMINI_MULTI_DOC", "6")))
    if len(files) > max_multi:
        print(
            f"[INFO] [GEMINI-BATCH] {len(files)} file(s) exceeds single-shot cap "
            f"({max_multi}); processing in sequential sub-batches"
        )
        all_results: List[dict] = []
        usages: list = []
        for i in range(0, len(files), max_multi):
            chunk = files[i : i + max_multi]
            chunk_results, chunk_usage = process_gastos_batch_with_gemini(
                chunk,
                max_retries=max_retries,
                concepto_catalog=concepto_catalog,
                tipo_de_pago_catalog=tipo_de_pago_catalog,
                concepto_document_comment=concepto_document_comment,
                tipo_de_pago_document_comment=tipo_de_pago_document_comment,
                business_rules=business_rules,
                tipo_de_gasto_context=tipo_de_gasto_context,
                tipo_de_gasto_document_comment=tipo_de_gasto_document_comment,
                model_name=model_id,
                spend_ctx=spend_ctx,
            )
            all_results.extend(chunk_results)
            if chunk_usage:
                usages.append(chunk_usage)
        return all_results, merge_usage_records(usages)

    if not GEMINI_API_KEY:
        print("[ERROR] [GEMINI-BATCH] Cannot process batch: GEMINI_API_KEY not configured")
        return (
            [_empty_gastos_extracted(fn, descripcion="ERROR: API key not configured")
             for _, fn in files],
            None,
        )

    # Render each file to its page-image list up front so we can fail fast on
    # unreadable files and so the per-document prompt knows the page counts.
    per_file_images: List[list] = []
    for content, filename in files:
        images = load_file_as_images(content, filename)
        per_file_images.append(images)
        if not images:
            print(
                f"[WARNING] [GEMINI-BATCH] '{filename}': produced 0 page images "
                f"({len(content)} bytes) - this document will yield an empty "
                f"result even though the batch call itself may succeed."
            )

    total_images = sum(len(p) for p in per_file_images)
    print(
        f"[DEBUG] [GEMINI-BATCH] Loaded {total_images} total page image(s) "
        f"across {len(files)} document(s)"
    )

    # Describe each document in the batch (filename + page count) so the model
    # understands when several images belong to the same multi-page PDF.
    doc_summary_lines = []
    for i, ((_, filename), pages) in enumerate(zip(files, per_file_images), start=1):
        doc_summary_lines.append(
            f"  - Documento #{i}: {filename} ({len(pages)} pagina(s))"
        )

    catalog_block = _build_gastos_catalog_prompt_block(
        concepto_catalog or [], tipo_de_pago_catalog or [],
        concepto_document_comment=concepto_document_comment,
        tipo_de_pago_document_comment=tipo_de_pago_document_comment,
    )
    catalog_block += _build_gastos_business_rules_prompt_block(business_rules or [])
    catalog_block += _build_gastos_tipo_de_gasto_context_block(
        tipo_de_gasto_context or [], document_comment=tipo_de_gasto_document_comment
    )
    batch_keys = (
        "nombre, documento, ncf, ncf_afectado, tipo_de_suplidor, tipo_de_gasto, "
        "descripcion, fecha, monto_en_servicios, monto_en_bienes, itbis, "
        "selectivo, descuento, propina, moneda, metodo_de_pago, campos_dudosos, "
        "razones_campos_dudosos, score"
    )
    if concepto_catalog:
        batch_keys += ", concepto"
    if tipo_de_pago_catalog:
        batch_keys += ", tipo_de_pago_erp"

    batch_prompt = (
        f"{GASTOS_SYSTEM_PROMPT}{catalog_block}\n\n"
        f"IMPORTANTE - Procesamiento por lotes: Te voy a enviar {len(files)} "
        f"documentos de recibos/facturas. Algunos pueden ser PDFs con varias "
        f"paginas; trata cada documento como UNA unidad y retorna UN solo "
        f"objeto JSON por documento.\n\n"
        f"Documentos en este lote:\n" + "\n".join(doc_summary_lines) + "\n\n"
        f"Debes retornar un ARRAY JSON con EXACTAMENTE {len(files)} elementos, "
        f"uno por cada documento, en el MISMO ORDEN en que se envian.\n\n"
        f"Cada elemento del array debe ser un objeto JSON con las claves: "
        f"{batch_keys}.\n\n"
        f"Retorna SOLO el array JSON, sin texto adicional, sin markdown, sin "
        f"explicaciones."
    )

    last_error: Optional[Exception] = None

    for attempt in range(max_retries):
        try:
            # Build the content list: prompt + per-document blocks. Each
            # document gets a header followed by ALL its page images.
            content_parts: list = [batch_prompt]
            for i, ((_, filename), pages) in enumerate(
                zip(files, per_file_images), start=1
            ):
                content_parts.append(
                    f"\n--- Documento #{i}: {filename} "
                    f"({len(pages)} pagina(s)) ---"
                )
                for page_idx, page_image in enumerate(pages, start=1):
                    if len(pages) > 1:
                        content_parts.append(
                            f"Pagina {page_idx} de {len(pages)}:"
                        )
                    content_parts.append(page_image)

            image_part_count = sum(1 for p in content_parts if not isinstance(p, str))
            print(
                f"[INFO] [GEMINI-BATCH] Calling Gemini (attempt {attempt + 1}/"
                f"{max_retries}, model={model_id}, content_parts="
                f"{len(content_parts)}, image_parts={image_part_count}, "
                f"prompt_chars={len(batch_prompt)})"
            )

            # ~400–500 tokens/doc for full receipt JSON; leave headroom.
            max_out = min(16384, max(4096, 768 * len(files)))
            # Large image batches need more wall-clock than the 180s default.
            timeout_s = max(180, 45 * max(1, len(files)))
            ctx = spend_ctx or {}
            level = ctx.get("thinking_level") or resolve_thinking_level(
                None, model_id
            )
            response = generate_inference_content(
                model_id,
                content_parts,
                thinking_level=level,
                max_output_tokens=max_out,
                temperature=0.1,
                timeout_s=timeout_s,
            )

            if not response or not getattr(response, "text", None):
                raise ValueError("Empty response from Gemini API")

            usage_record = record_usage_from_response(
                response,
                model=model_id,
                source="gastos_batch",
                thinking_level=level,
                organization_id=ctx.get("organization_id"),
                user_id=ctx.get("user_id"),
                client_id=ctx.get("client_id"),
                metadata={
                    "file_count": len(files),
                    "filenames": [fn for _, fn in files],
                },
            )

            parsed = _parse_json_loose(response.text)

            if isinstance(parsed, dict):
                parsed = [parsed]

            if not isinstance(parsed, list):
                raise ValueError(f"Expected JSON array from batch, got {type(parsed).__name__}")

            # Truncated JSON (common on large batches) used to be padded with
            # empty rows → every missing slot became "Reintentar". Treat a
            # shortfall as a retryable failure so we fall back to parallel.
            if len(parsed) != len(files):
                raise ValueError(
                    f"Incomplete batch: expected {len(files)} results, "
                    f"Gemini returned {len(parsed)}"
                )

            results: List[dict] = []
            for i, (_, filename) in enumerate(files):
                if isinstance(parsed[i], dict):
                    normalized = _normalize_gastos_extracted(
                        parsed[i], filename,
                        concepto_catalog=concepto_catalog,
                        tipo_de_pago_catalog=tipo_de_pago_catalog,
                    )
                    is_empty = not any(
                        normalized.get(k) for k in
                        ("nombre", "documento", "ncf", "tipo_de_suplidor", "fecha")
                    )
                    if is_empty:
                        print(
                            f"[WARNING] [GEMINI-BATCH] '{filename}': result #{i} is "
                            f"essentially empty (score={normalized.get('score')}) - "
                            f"check if its {len(per_file_images[i])} page image(s) "
                            f"were readable."
                        )
                    results.append(normalized)
                else:
                    raise ValueError(
                        f"Incomplete batch: result #{i} for '{filename}' "
                        f"is {type(parsed[i]).__name__}, expected object"
                    )
            print(f"[INFO] [GEMINI-BATCH] Batch completed: {len(results)} result(s)")
            return results, usage_record

        except Exception as e:
            last_error = e
            print(
                f"[ERROR] [GEMINI-BATCH] Attempt {attempt + 1}/{max_retries} "
                f"failed: {e}"
            )
            if _is_retryable_error(e) and attempt < max_retries - 1:
                print("[INFO] [GEMINI-BATCH] Retryable error, backing off before retry")
                time.sleep(2 ** attempt)
            elif not _is_retryable_error(e):
                break

    print(f"[ERROR] [GEMINI-BATCH] Gemini batch processing failed: {last_error}")
    # Fall back to parallel per-file so a bad batch parse doesn't serialize N calls.
    print(
        f"[INFO] [GEMINI-BATCH] Falling back to parallel per-file processing "
        f"for {len(files)} files"
    )
    return _process_gastos_files_parallel(files, **parallel_kwargs)


def _fill_gastos_template_xls(rows: list, tax_column_mapping: Optional[dict] = None) -> Path:
    """
    Write GASTOS_OUTPUT_XLS by filling the official Carga Masiva template with
    the given (already normalized) receipt rows. The template's dropdowns, named
    ranges and Nomencladores sheet are preserved verbatim (see xls_template).

    tax_column_mapping: optional per-client {itbis|selectivo|descuento|
    propina: 1..5} (see composables/useClientTaxColumnMapping.ts) deciding
    which "Impuesto N" column each amount is written into. Falls back to
    GASTOS_DEFAULT_TAX_COLUMN_MAPPING (itbis -> Impuesto 1, selectivo -> Impuesto 2)
    when not provided, matching the export's original behavior.
    """
    if not GASTOS_TEMPLATE_XLS_SOURCE.exists():
        raise FileNotFoundError(f"Template file not found: {GASTOS_TEMPLATE_XLS_SOURCE}")
    prepared = []
    for r in rows:
        row = prepare_gastos_export_row(r)
        row.update(resolve_gastos_tax_columns(row, tax_column_mapping))
        prepared.append(row)
    return fill_gastos_xls_template(
        GASTOS_TEMPLATE_XLS_SOURCE,
        GASTOS_OUTPUT_XLS,
        prepared,
        GASTOS_EXCEL_FIELD_MAPPINGS,
        GASTOS_EXCEL_TEXT_FIELDS,
        GASTOS_EXCEL_NUMERIC_FIELDS,
        GASTOS_EXCEL_INT_FIELDS,
        date_fields=GASTOS_EXCEL_DATE_FIELDS,
    )


def populate_gastos_excel_template(data: dict, tax_column_mapping: Optional[dict] = None):
    """Fill the gastos template with a single extracted receipt (fresh export)."""
    try:
        return _fill_gastos_template_xls([data], tax_column_mapping=tax_column_mapping)
    except Exception as e:
        print(f"[ERROR] Excel processing failed: {e}")
        raise


@app.get("/")
async def root():
    """Health check endpoint"""
    return {"status": "ok", "message": "Receipt Processing API is running"}


@app.get("/thinking-levels")
async def get_thinking_levels():
    """
    Return the thinking-speed → Gemini model map from server ENV.

    Used by the UI selector so frontend model ids stay in sync with
    THINKING_LEVEL_*_MODEL without hardcoding them in the client.
    """
    models = get_thinking_level_models()
    return {
        "default": "moderado",
        "models": models,
        "levels": [
            {"value": level, "model": models[level]}
            for level in ("rapido", "moderado", "profundo")
            if level in models
        ],
    }


@app.get("/models")
async def get_models(generate_content_only: bool = True):
    """
    List Gemini models available to the configured API key.

    Open in the browser: http://127.0.0.1:8000/models
    Pass ?generate_content_only=false to include models that cannot run
    generateContent (e.g. embedding-only models).
    """
    if not GEMINI_API_KEY:
        raise HTTPException(
            status_code=503,
            detail="GEMINI_API_KEY is not configured on the server.",
        )
    try:
        models = await asyncio.to_thread(
            list_available_gemini_models,
            generate_content_only=generate_content_only,
        )
    except Exception as e:
        print(f"[ERROR] [/models] Failed to list Gemini models: {e}")
        raise HTTPException(
            status_code=502,
            detail=f"Failed to list Gemini models: {e}",
        )

    return {
        "count": len(models),
        "configured": {
            "individual": INDIVIDUAL_MODEL,
            "batch": BATCH_MODEL,
        },
        "models": models,
    }


@app.post("/upload")
async def upload_file(
    file: UploadFile = File(...),
    concepto_catalog: Optional[str] = Form(None),
    tipo_de_pago_catalog: Optional[str] = Form(None),
    concepto_document_comment: Optional[str] = Form(""),
    tipo_de_pago_document_comment: Optional[str] = Form(""),
    business_rules: Optional[str] = Form(None),
    tipo_de_gasto_context: Optional[str] = Form(None),
    tipo_de_gasto_document_comment: Optional[str] = Form(""),
    tax_column_mapping: Optional[str] = Form(None),
    model: Optional[str] = Form(None),
    thinking_level: Optional[str] = Form(None),
    user_id: Optional[str] = Form(None),
    organization_id: Optional[str] = Form(None),
    client_id: Optional[str] = Form(None),
):
    """
    Upload and process a receipt/invoice file (supports: PDF, PNG, JPG, JPEG).

    model: optional Gemini/Gemma model id from the UI thinking-level selector.
    Whitelisted values only; falls back to INDIVIDUAL_MODEL.

    thinking_level / user_id / organization_id / client_id: optional spend
    attribution. When organization_id + user_id are set (and Supabase ENV is
    configured), token usage is persisted to api_token_usage.

    concepto_catalog / tipo_de_pago_catalog: optional JSON-encoded arrays of
    {document_type, document_id, description} for the client currently being
    scanned (see client_documents / document_attributes). When provided, the
    LLM is asked to classify the document into one of these ERP-specific
    options and the matching ERP id is written to Concepto Id / Tipo de Pago
    Id in the export.

    concepto_document_comment / tipo_de_pago_document_comment: optional
    document-level comments (the "comment" column on the client_documents
    container itself, e.g. its "Gastos" group) that give the LLM extra
    context on top of each catalog option's own description. Sent as "" when
    the document has no comment, in which case behavior is unchanged.

    business_rules: optional JSON-encoded array of {rule_type, rule_value,
    description} for the client currently being scanned (see
    client_business_rules / business_rule_attributes). Unlike the catalogs
    above, these are free-form context (not classification options) that
    help the LLM make better decisions for this specific client.

    tipo_de_gasto_context / tipo_de_gasto_document_comment: optional
    JSON-encoded array of {document_type, document_id, description} (and its
    document-level comment) for a client_documents container the user picked
    specifically to give the LLM context for THIS client when choosing among
    the FIXED tipo_de_gasto options - it never introduces new tipo_de_gasto
    values, only helps pick among the existing 11.

    tax_column_mapping: optional JSON-encoded {itbis|selectivo|descuento|
    propina: 1..5} for the client currently being scanned (see
    client_tax_column_mappings), deciding which "Impuesto N" column of the
    export each amount is written into. Falls back to the original
    itbis -> Impuesto 1 / selectivo -> Impuesto 2 behavior when omitted.
    """
    try:
        allowed_extensions = {'.pdf', '.png', '.jpg', '.jpeg'}
        file_extension = Path(file.filename).suffix.lower() if file.filename else ''
        
        if file_extension not in allowed_extensions:
            raise HTTPException(
                status_code=400, 
                detail=f"Unsupported file type. Supported formats: PDF, PNG, JPG, JPEG."
            )
        
        file_content = await file.read()
        file_hash = calculate_file_hash(file_content)
        level = resolve_thinking_level(thinking_level, model)
        model_id = resolve_inference_model(
            model, default=INDIVIDUAL_MODEL, thinking_level=level,
        )
        spend_ctx = {
            "thinking_level": level,
            "user_id": (user_id or "").strip() or None,
            "organization_id": (organization_id or "").strip() or None,
            "client_id": (client_id or "").strip() or None,
        }
        print(
            f"[INFO] [/upload] Received '{file.filename}' "
            f"({len(file_content)} bytes, content_type={file.content_type}, "
            f"hash={file_hash[:8]}, model={model_id}, level={level})"
        )

        # NOTE: no duplicate detection here - this endpoint processes a
        # single file with no siblings to compare against in the same
        # request, and checking against persisted history would incorrectly
        # block legitimate re-processing (e.g. retry/reevaluate, or
        # re-scanning after a page reload). See /upload-batch for
        # within-batch duplicate detection.

        concepto_list = _clean_catalog(_parse_catalog_param(concepto_catalog, "concepto_catalog"))
        tipo_de_pago_list = _clean_catalog(
            _parse_catalog_param(tipo_de_pago_catalog, "tipo_de_pago_catalog")
        )
        concepto_comment = (concepto_document_comment or "").strip()
        tipo_de_pago_comment = (tipo_de_pago_document_comment or "").strip()
        business_rules_list = _clean_business_rules(
            _parse_catalog_param(business_rules, "business_rules")
        )
        tipo_de_gasto_context_list = _clean_document_context(
            _parse_catalog_param(tipo_de_gasto_context, "tipo_de_gasto_context")
        )
        tipo_de_gasto_comment = (tipo_de_gasto_document_comment or "").strip()
        tax_column_mapping_dict = _parse_tax_column_mapping_param(tax_column_mapping)

        # Process with Gemini (rate limited to 5 concurrent)
        async with gemini_semaphore:
            extracted_data, usage_record = await asyncio.to_thread(
                process_gastos_with_gemini, file_content, file.filename,
                concepto_catalog=concepto_list,
                tipo_de_pago_catalog=tipo_de_pago_list,
                concepto_document_comment=concepto_comment,
                tipo_de_pago_document_comment=tipo_de_pago_comment,
                business_rules=business_rules_list,
                tipo_de_gasto_context=tipo_de_gasto_context_list,
                tipo_de_gasto_document_comment=tipo_de_gasto_comment,
                model_name=model_id,
                spend_ctx=spend_ctx,
            )
        
        print(f"[INFO] [/upload] '{file.filename}' processed, score={extracted_data.get('score')}")

        populate_gastos_excel_template(extracted_data, tax_column_mapping=tax_column_mapping_dict)
        
        # Save to history
        history = load_history()
        history.append({
            "hash": file_hash,
            "filename": file.filename,
            "processed_at": datetime.now().isoformat(),
            "data": extracted_data
        })
        save_history(history)
        
        return {
            "status": "success",
            "message": f"File {file.filename} processed successfully",
            "data": extracted_data,
            "hash": file_hash,
            "usage": usage_record,
        }
        
    except HTTPException:
        raise
    except Exception as e:
        print(f"[ERROR] Upload failed for {file.filename}: {e}")
        raise HTTPException(status_code=500, detail=f"Error processing file: {str(e)}")


@app.post("/upload-batch")
async def upload_batch(
    files: List[UploadFile] = File(...),
    concepto_catalog: Optional[str] = Form(None),
    tipo_de_pago_catalog: Optional[str] = Form(None),
    concepto_document_comment: Optional[str] = Form(""),
    tipo_de_pago_document_comment: Optional[str] = Form(""),
    business_rules: Optional[str] = Form(None),
    tipo_de_gasto_context: Optional[str] = Form(None),
    tipo_de_gasto_document_comment: Optional[str] = Form(""),
    tax_column_mapping: Optional[str] = Form(None),
    model: Optional[str] = Form(None),
    thinking_level: Optional[str] = Form(None),
    user_id: Optional[str] = Form(None),
    organization_id: Optional[str] = Form(None),
    client_id: Optional[str] = Form(None),
):
    """
    Upload and process MULTIPLE receipt/invoice files in a SINGLE Gemini API call.

    Accepts multiple files (chunked server-side into groups of ~6 for the
    Gemini multi-doc call). Supported formats: PNG / JPG / JPEG / PDF. PDFs are
    rasterized to images (one image per page, capped at PDF_MAX_PAGES) and all
    pages of a PDF are still represented as ONE entry in the response.

    model: optional Gemini/Gemma model id from the UI thinking-level selector.
    Whitelisted values only; falls back to BATCH_MODEL.

    thinking_level / user_id / organization_id / client_id: optional spend
    attribution for api_token_usage persistence.

    concepto_catalog / tipo_de_pago_catalog: optional JSON-encoded arrays of
    {document_type, document_id, description} for the client currently being
    scanned. Applied to every file in the batch.

    concepto_document_comment / tipo_de_pago_document_comment: optional
    document-level comments applied to every file in the batch, sent as ""
    when the document has no comment (behavior stays unchanged).

    business_rules: optional JSON-encoded array of {rule_type, rule_value,
    description} for the client currently being scanned (see
    client_business_rules / business_rule_attributes), applied to every file
    in the batch as free-form context for the LLM.

    tipo_de_gasto_context / tipo_de_gasto_document_comment: optional
    JSON-encoded array of {document_type, document_id, description} (and its
    document-level comment) for a client_documents container picked
    specifically to help the LLM choose among the FIXED tipo_de_gasto
    options for every file in the batch - never introduces new values.

    tax_column_mapping: optional JSON-encoded {itbis|selectivo|descuento|
    propina: 1..5} for the client currently being scanned (see
    client_tax_column_mappings), applied to every file in the batch when
    writing its amounts into the export's "Impuesto N" columns.

    Duplicate detection: files are compared against OTHER FILES IN THIS SAME
    BATCH only (never against previously processed history, so re-scanning
    the same document later - e.g. after a page reload - is never blocked).
    A file is flagged "duplicate" (and excluded from the Excel export /
    history) when either (a) its bytes are identical to another file in the
    batch (exact hash match - the Gemini call is skipped entirely), or (b)
    its extracted data matches another receipt in the batch by
    NCF+documento (or, when no NCF is present, by documento/nombre + fecha +
    montos).

    Returns:
        {
            "status": "success",
            "results": [
                {"status": "success",   "filename": "...", "hash": "...", "data": {...}},
                {"status": "duplicate", "filename": "...", "message": "...", "duplicate_of": "...", "data": {...}},
                {"status": "error",     "filename": "...", "message": "..."},
                ...
            ]
        }
    Results are in the same order as the uploaded files.
    """
    if not files:
        raise HTTPException(status_code=400, detail="No files provided")

    model_id = resolve_inference_model(
        model, default=BATCH_MODEL, thinking_level=thinking_level,
    )
    level = resolve_thinking_level(thinking_level, model_id)
    spend_ctx = {
        "thinking_level": level,
        "user_id": (user_id or "").strip() or None,
        "organization_id": (organization_id or "").strip() or None,
        "client_id": (client_id or "").strip() or None,
    }
    print(
        f"[INFO] [/upload-batch] Received request with {len(files)} file(s), "
        f"model={model_id}, level={level}, "
        f"org={'yes' if spend_ctx.get('organization_id') else 'no'}, "
        f"user={'yes' if spend_ctx.get('user_id') else 'no'}, "
        f"client={'yes' if spend_ctx.get('client_id') else 'no'}"
    )

    allowed_extensions = {'.pdf', '.png', '.jpg', '.jpeg'}

    # Duplicate-detection state, scoped to THIS batch only (not persisted
    # history - see _new_dedupe_state), extended below as we walk the
    # batch so exact file duplicates are caught WITHOUT wasting a Gemini
    # call (e.g. the same receipt picked twice, or two files with identical
    # bytes but different names, such as duplicate pages in a multi-page
    # PDF that got split client-side).
    seen_hashes, hash_to_filename, seen_keys, key_to_filename = _new_dedupe_state()

    # Read each upload and classify it. Pending entries will be sent to Gemini.
    entries: list = []
    for upload in files:
        filename = upload.filename or "unknown"
        ext = Path(filename).suffix.lower()

        if ext not in allowed_extensions:
            print(f"[WARNING] [/upload-batch] '{filename}': unsupported extension '{ext}', skipping")
            entries.append({
                "filename": filename,
                "status": "error",
                "message": f"Unsupported file type: {ext or 'unknown'}",
            })
            continue

        try:
            content = await upload.read()
        except Exception as e:
            print(f"[ERROR] [/upload-batch] '{filename}': failed to read upload body: {e}")
            entries.append({
                "filename": filename,
                "status": "error",
                "message": f"Failed to read upload: {e}",
            })
            continue

        print(
            f"[INFO] [/upload-batch] '{filename}': read {len(content)} bytes "
            f"(content_type={upload.content_type})"
        )
        if len(content) == 0:
            print(
                f"[WARNING] [/upload-batch] '{filename}': uploaded file is 0 "
                f"bytes - the browser may have sent an empty body for this file."
            )

        file_hash = calculate_file_hash(content)
        if file_hash in seen_hashes:
            duplicate_of = hash_to_filename.get(file_hash, "?")
            print(
                f"[INFO] [/upload-batch] '{filename}': duplicate of "
                f"'{duplicate_of}' (identical file content) - skipping "
                f"Gemini call"
            )
            entries.append({
                "filename": filename,
                "status": "duplicate",
                "message": f"Archivo idéntico a '{duplicate_of}' (mismo contenido).",
                "duplicate_of": duplicate_of,
            })
            continue

        seen_hashes.add(file_hash)
        hash_to_filename[file_hash] = filename

        entries.append({
            "filename": filename,
            "content": content,
            "hash": file_hash,
            "status": "pending",
        })

    batch_inputs: List[Tuple[bytes, str]] = [
        (e["content"], e["filename"]) for e in entries if e["status"] == "pending"
    ]

    print(
        f"[INFO] [/upload-batch] {len(batch_inputs)}/{len(files)} file(s) "
        f"eligible for Gemini processing"
    )

    concepto_list = _clean_catalog(_parse_catalog_param(concepto_catalog, "concepto_catalog"))
    tipo_de_pago_list = _clean_catalog(
        _parse_catalog_param(tipo_de_pago_catalog, "tipo_de_pago_catalog")
    )
    concepto_comment = (concepto_document_comment or "").strip()
    tipo_de_pago_comment = (tipo_de_pago_document_comment or "").strip()
    business_rules_list = _clean_business_rules(
        _parse_catalog_param(business_rules, "business_rules")
    )
    tipo_de_gasto_context_list = _clean_document_context(
        _parse_catalog_param(tipo_de_gasto_context, "tipo_de_gasto_context")
    )
    tipo_de_gasto_comment = (tipo_de_gasto_document_comment or "").strip()
    tax_column_mapping_dict = _parse_tax_column_mapping_param(tax_column_mapping)

    extracted_results: List[dict] = []
    usage_record: Optional[dict] = None
    if batch_inputs:
        # Reuse the same semaphore to coordinate concurrency with single uploads.
        async with gemini_semaphore:
            extracted_results, usage_record = await asyncio.to_thread(
                process_gastos_batch_with_gemini, batch_inputs,
                concepto_catalog=concepto_list,
                tipo_de_pago_catalog=tipo_de_pago_list,
                concepto_document_comment=concepto_comment,
                tipo_de_pago_document_comment=tipo_de_pago_comment,
                business_rules=business_rules_list,
                tipo_de_gasto_context=tipo_de_gasto_context_list,
                tipo_de_gasto_document_comment=tipo_de_gasto_comment,
                model_name=model_id,
                spend_ctx=spend_ctx,
            )

    history = load_history()
    response_results: list = []
    extracted_idx = 0

    for entry in entries:
        if entry["status"] == "pending":
            if extracted_idx < len(extracted_results):
                data = extracted_results[extracted_idx]
            else:
                data = _empty_gastos_extracted(entry["filename"])
            extracted_idx += 1

            # Duplicate check: different file bytes, but the extracted data
            # (NCF/documento or fecha+montos) matches a receipt already
            # counted - either earlier in this same batch or in a previous
            # session. Exclude it from the export/history so it isn't
            # double-counted, without failing the whole batch.
            receipt_key = _gastos_receipt_identity_key(data)
            if receipt_key and receipt_key in seen_keys:
                duplicate_of = key_to_filename.get(receipt_key, "?")
                print(
                    f"[INFO] [/upload-batch] '{entry['filename']}': duplicate "
                    f"of '{duplicate_of}' (matching extracted receipt data) - "
                    f"excluding from export"
                )
                response_results.append({
                    "status": "duplicate",
                    "filename": entry["filename"],
                    "hash": entry["hash"],
                    "message": f"Mismos datos que '{duplicate_of}' (posible recibo duplicado).",
                    "duplicate_of": duplicate_of,
                    "data": data,
                })
                continue

            if receipt_key:
                seen_keys.add(receipt_key)
                key_to_filename[receipt_key] = entry["filename"]

            try:
                populate_gastos_excel_template(data, tax_column_mapping=tax_column_mapping_dict)
            except Exception as e:
                print(f"[ERROR] Excel populate failed for {entry['filename']}: {e}")

            history.append({
                "hash": entry["hash"],
                "filename": entry["filename"],
                "processed_at": datetime.now().isoformat(),
                "data": data,
            })

            response_results.append({
                "status": "success",
                "filename": entry["filename"],
                "hash": entry["hash"],
                "data": data,
            })
        else:
            result_entry = {
                "status": entry["status"],
                "filename": entry["filename"],
                "message": entry.get("message", "Unknown error"),
            }
            if "duplicate_of" in entry:
                result_entry["duplicate_of"] = entry["duplicate_of"]
            response_results.append(result_entry)

    save_history(history)

    print(
        f"[INFO] [/upload-batch] Responding with {len(response_results)} "
    )

    return {
        "status": "success",
        "count": len(response_results),
        "results": response_results,
        "usage": usage_record,
    }


def regenerate_gastos_excel_from_data(files_data: list, tax_column_mapping: Optional[dict] = None):
    """Regenerate the gastos export by filling the template with edited data.

    Score is excluded from export. The template is filled (not recreated) so
    its dropdowns / data validations remain intact for the destination system.
    """
    try:
        return _fill_gastos_template_xls(files_data or [], tax_column_mapping=tax_column_mapping)
    except Exception as e:
        print(f"[ERROR] Regenerating Excel failed: {e}")
        raise


@app.post("/download")
async def download_excel_post(payload: Optional[Any] = Body(None)):
    """
    Download Excel file as .xls. If a body is provided, regenerates with
    edited data first.

    Accepts either the legacy shape (a bare JSON array of row dicts) or
    {"files_data": [...], "tax_column_mapping": {...}} where
    tax_column_mapping is the client's {itbis|selectivo|descuento|propina:
    1..5} Impuesto-column mapping (see composables/useClientTaxColumnMapping.ts).
    """
    files_data: Optional[list] = None
    tax_column_mapping: Optional[dict] = None
    if isinstance(payload, list):
        files_data = payload
    elif isinstance(payload, dict):
        files_data = payload.get("files_data")
        tax_column_mapping = payload.get("tax_column_mapping")

    if files_data:
        regenerate_gastos_excel_from_data(files_data, tax_column_mapping=tax_column_mapping)
    
    if not GASTOS_OUTPUT_XLS.exists():
        if GASTOS_OUTPUT_FILE.exists():
            convert_xlsx_to_xls(GASTOS_OUTPUT_FILE, GASTOS_OUTPUT_XLS)
        else:
            raise HTTPException(status_code=404, detail="No processed file available")
    
    return FileResponse(
        path=GASTOS_OUTPUT_XLS,
        filename="processed_receipts.xls",
        media_type="application/vnd.ms-excel",
    )


@app.get("/download")
async def download_excel_get():
    """Download the processed Excel file as .xls."""
    if not GASTOS_OUTPUT_XLS.exists():
        if GASTOS_OUTPUT_FILE.exists():
            convert_xlsx_to_xls(GASTOS_OUTPUT_FILE, GASTOS_OUTPUT_XLS)
        else:
            raise HTTPException(status_code=404, detail="No processed file available")
    
    return FileResponse(
        path=GASTOS_OUTPUT_XLS,
        filename="processed_receipts.xls",
        media_type="application/vnd.ms-excel",
    )


# ---------------------------------------------------------------------------
# Suplidores routes (scan + export) — defined in suplidores_server.py
# ---------------------------------------------------------------------------
from suplidores_server import router as suplidores_router  # noqa: E402
app.include_router(suplidores_router)


if __name__ == "__main__":
    import uvicorn

    host = os.getenv("HOST", "127.0.0.1")
    port = int(os.getenv("PORT", "8000"))
    reload_enabled = os.getenv("RELOAD", "true").lower() != "false"

    print(f"[INFO] Starting server on {host}:{port} (reload={reload_enabled})")

    if reload_enabled:
        uvicorn.run("server:app", host=host, port=port, reload=True, reload_dirs=[str(BASE_DIR)])
    else:
        uvicorn.run(app, host=host, port=port, reload=False)

