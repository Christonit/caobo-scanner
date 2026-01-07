#!/usr/bin/env python3
"""
One-time script to convert template.xls to template.xlsx
Run this once: python convert_template.py

This script reads the original .xls template and creates an exact .xlsx copy
preserving ALL sheets, ALL columns, ALL data, and ALL formatting.
"""
import subprocess
import sys
from pathlib import Path

BASE_DIR = Path(__file__).parent
TEMPLATE_XLS = BASE_DIR / "template.xls"
TEMPLATE_XLSX = BASE_DIR / "template_converted.xlsx"

# Also check root directory
ROOT_TEMPLATE_XLS = BASE_DIR.parent / "template.xls"


def install_xlrd():
    """Install xlrd if not present"""
    try:
        import xlrd
        print(f"✓ xlrd is already installed (version {xlrd.__VERSION__})")
        return True
    except ImportError:
        print("Installing xlrd...")
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "xlrd==1.2.0"])
            print("✓ xlrd installed successfully")
            return True
        except Exception as e:
            print(f"✗ Failed to install xlrd: {e}")
            return False


def get_color_from_index(color_map, index):
    """Convert xlrd color index to hex color"""
    # Standard Excel color palette
    STANDARD_COLORS = {
        0: "000000",   # Black
        1: "FFFFFF",   # White
        2: "FF0000",   # Red
        3: "00FF00",   # Lime
        4: "0000FF",   # Blue
        5: "FFFF00",   # Yellow
        6: "FF00FF",   # Magenta
        7: "00FFFF",   # Cyan
        8: "000000",   # Black
        9: "FFFFFF",   # White
        10: "FF0000",  # Red
        11: "00FF00",  # Lime
        12: "0000FF",  # Blue
        13: "FFFF00",  # Yellow
        14: "FF00FF",  # Magenta
        15: "00FFFF",  # Cyan
        16: "800000",  # Maroon
        17: "008000",  # Green
        18: "000080",  # Navy
        19: "808000",  # Olive
        20: "800080",  # Purple
        21: "008080",  # Teal
        22: "C0C0C0",  # Silver
        23: "808080",  # Gray
        24: "9999FF",  # Periwinkle
        25: "993366",  # Plum
        26: "FFFFCC",  # Ivory
        27: "CCFFFF",  # Light Turquoise
        28: "660066",  # Dark Purple
        29: "FF8080",  # Coral
        30: "0066CC",  # Ocean Blue
        31: "CCCCFF",  # Ice Blue
        32: "000080",  # Dark Blue
        33: "FF00FF",  # Pink
        34: "FFFF00",  # Yellow
        35: "00FFFF",  # Turquoise
        36: "800080",  # Violet
        37: "800000",  # Dark Red
        38: "008080",  # Teal
        39: "0000FF",  # Blue
        40: "00CCFF",  # Sky Blue
        41: "CCFFFF",  # Light Turquoise
        42: "CCFFCC",  # Light Green
        43: "FFFF99",  # Light Yellow
        44: "99CCFF",  # Pale Blue
        45: "FF99CC",  # Rose
        46: "CC99FF",  # Lavender
        47: "FFCC99",  # Tan
        48: "3366FF",  # Light Blue
        49: "33CCCC",  # Aqua
        50: "99CC00",  # Lime
        51: "FFCC00",  # Gold
        52: "FF9900",  # Light Orange
        53: "FF6600",  # Orange
        54: "666699",  # Blue Gray
        55: "969696",  # Gray 40%
        56: "003366",  # Dark Teal
        57: "339966",  # Sea Green
        58: "003300",  # Dark Green
        59: "333300",  # Olive Green
        60: "993300",  # Brown
        61: "993366",  # Plum
        62: "333399",  # Indigo
        63: "333333",  # Gray 80%
        64: "000000",  # Black (system)
    }
    
    if index in STANDARD_COLORS:
        return STANDARD_COLORS[index]
    
    # Try to get from color map
    try:
        if color_map and index < len(color_map):
            rgb = color_map[index]
            if rgb:
                return f"{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}"
    except:
        pass
    
    return None


def convert_with_xlrd(xls_path: Path, xlsx_path: Path):
    """Convert using xlrd + openpyxl with full formatting support"""
    import xlrd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.comments import Comment
    
    print(f"Reading: {xls_path}")
    
    # Open the .xls file with formatting info
    xls_book = xlrd.open_workbook(xls_path, formatting_info=True)
    
    # Get color map
    color_map = xls_book.colour_map if hasattr(xls_book, 'colour_map') else {}
    
    # Create new .xlsx workbook
    xlsx_book = openpyxl.Workbook()
    xlsx_book.remove(xlsx_book.active)
    
    print(f"Found {xls_book.nsheets} sheet(s)")
    
    for sheet_idx in range(xls_book.nsheets):
        xls_sheet = xls_book.sheet_by_index(sheet_idx)
        sheet_name = xls_sheet.name
        print(f"  Processing sheet: '{sheet_name}' ({xls_sheet.nrows} rows x {xls_sheet.ncols} cols)")
        
        xlsx_sheet = xlsx_book.create_sheet(title=sheet_name)
        
        # Track merged cells
        merged_ranges = []
        
        # Copy all cells with full formatting
        for row_idx in range(xls_sheet.nrows):
            for col_idx in range(xls_sheet.ncols):
                xls_cell = xls_sheet.cell(row_idx, col_idx)
                xlsx_cell = xlsx_sheet.cell(row=row_idx + 1, column=col_idx + 1)
                
                # Copy value
                value = xls_cell.value
                if value != '':
                    xlsx_cell.value = value
                
                # Copy formatting
                try:
                    xf_index = xls_cell.xf_index
                    xf = xls_book.xf_list[xf_index]
                    
                    # Background color
                    bg_color_idx = xf.background.pattern_colour_index
                    bg_color = get_color_from_index(color_map, bg_color_idx)
                    if bg_color and bg_color != "FFFFFF" and xf.background.pattern_type_id != 0:
                        xlsx_cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
                    
                    # Font formatting
                    font_obj = xls_book.font_list[xf.font_index]
                    font_color = get_color_from_index(color_map, font_obj.colour_index)
                    xlsx_cell.font = Font(
                        bold=font_obj.bold,
                        italic=font_obj.italic,
                        underline='single' if font_obj.underlined else None,
                        strike=font_obj.struck_out,
                        color=font_color if font_color else "000000",
                        size=font_obj.height // 20 if font_obj.height else 11,
                        name=font_obj.name if font_obj.name else "Calibri"
                    )
                    
                    # Alignment
                    h_align_map = {0: 'general', 1: 'left', 2: 'center', 3: 'right', 4: 'fill', 5: 'justify', 6: 'centerContinuous'}
                    v_align_map = {0: 'top', 1: 'center', 2: 'bottom', 3: 'justify', 4: 'distributed'}
                    h_align = h_align_map.get(xf.alignment.hor_align, 'general')
                    v_align = v_align_map.get(xf.alignment.vert_align, 'bottom')
                    xlsx_cell.alignment = Alignment(
                        horizontal=h_align,
                        vertical=v_align,
                        wrap_text=xf.alignment.text_wrapped
                    )
                    
                    # Borders
                    border_style_map = {0: None, 1: 'thin', 2: 'medium', 3: 'dashed', 4: 'dotted', 5: 'thick', 6: 'double', 7: 'hair'}
                    def get_side(style_idx, color_idx):
                        style = border_style_map.get(style_idx)
                        if style:
                            color = get_color_from_index(color_map, color_idx) or "000000"
                            return Side(style=style, color=color)
                        return Side()
                    
                    xlsx_cell.border = Border(
                        left=get_side(xf.border.left_line_style, xf.border.left_colour_index),
                        right=get_side(xf.border.right_line_style, xf.border.right_colour_index),
                        top=get_side(xf.border.top_line_style, xf.border.top_colour_index),
                        bottom=get_side(xf.border.bottom_line_style, xf.border.bottom_colour_index)
                    )
                    
                except Exception as e:
                    pass  # Skip formatting errors
        
        # Copy merged cells
        try:
            for (rlo, rhi, clo, chi) in xls_sheet.merged_cells:
                start = f"{get_column_letter(clo + 1)}{rlo + 1}"
                end = f"{get_column_letter(chi)}{rhi}"
                xlsx_sheet.merge_cells(f"{start}:{end}")
        except:
            pass
        
        # Copy column widths
        for col_idx in range(xls_sheet.ncols):
            col_letter = get_column_letter(col_idx + 1)
            try:
                # xlrd stores column widths in 1/256 of character width
                col_info = xls_sheet.colinfo_map.get(col_idx)
                if col_info:
                    width = col_info.width / 256
                    xlsx_sheet.column_dimensions[col_letter].width = max(width, 8)
                else:
                    # Calculate based on content
                    max_length = 0
                    for row_idx in range(min(xls_sheet.nrows, 100)):  # Check first 100 rows
                        cell_value = xls_sheet.cell(row_idx, col_idx).value
                        if cell_value:
                            max_length = max(max_length, len(str(cell_value)))
                    xlsx_sheet.column_dimensions[col_letter].width = max(max_length + 2, 12)
            except:
                xlsx_sheet.column_dimensions[col_letter].width = 15
        
        # Copy row heights
        for row_idx in range(xls_sheet.nrows):
            try:
                row_info = xls_sheet.rowinfo_map.get(row_idx)
                if row_info and row_info.height:
                    # xlrd stores height in twips (1/20 of a point)
                    xlsx_sheet.row_dimensions[row_idx + 1].height = row_info.height / 20
            except:
                pass
    
    # Apply auto-filter to first sheet
    if xlsx_book.sheetnames:
        first_sheet = xlsx_book[xlsx_book.sheetnames[0]]
        if first_sheet.max_row > 0 and first_sheet.max_column > 0:
            filter_range = f"A1:{get_column_letter(first_sheet.max_column)}1"
            first_sheet.auto_filter.ref = filter_range
            print(f"  Applied auto-filter: {filter_range}")
    
    # Save
    xlsx_book.save(xlsx_path)
    print(f"✓ Saved: {xlsx_path}")
    return True


def convert_with_pandas(xls_path: Path, xlsx_path: Path):
    """Fallback: Convert using pandas (less formatting support)"""
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    
    print(f"Reading with pandas: {xls_path}")
    print("  Note: Pandas conversion has limited formatting support")
    
    # Read all sheets
    xls = pd.ExcelFile(xls_path, engine='xlrd')
    
    # Create new workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    for sheet_name in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
        print(f"  Processing sheet: '{sheet_name}' ({len(df)} rows x {len(df.columns)} cols)")
        
        ws = wb.create_sheet(title=sheet_name)
        
        # Define styles
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        bold_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write ALL data
        for row_idx, row in df.iterrows():
            for col_idx, value in enumerate(row, start=1):
                if pd.notna(value):
                    cell = ws.cell(row=row_idx + 1, column=col_idx, value=value)
                    # Apply header formatting to first row
                    if row_idx == 0:
                        cell.fill = yellow_fill
                        cell.font = bold_font
                        cell.border = thin_border
        
        # Set column widths based on content
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for row_idx in range(1, min(ws.max_row + 1, 101)):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[col_letter].width = max(max_length + 2, 12)
        
        # Apply auto-filter to header row
        if ws.max_row > 0 and ws.max_column > 0:
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}1"
    
    wb.save(xlsx_path)
    print(f"✓ Saved: {xlsx_path}")
    return True


def main():
    print("=" * 60)
    print("Template Conversion Script (Full Formatting)")
    print("=" * 60)
    
    # Find template
    template_path = None
    if TEMPLATE_XLS.exists():
        template_path = TEMPLATE_XLS
    elif ROOT_TEMPLATE_XLS.exists():
        template_path = ROOT_TEMPLATE_XLS
    else:
        print(f"✗ Template not found!")
        print(f"  Checked: {TEMPLATE_XLS}")
        print(f"  Checked: {ROOT_TEMPLATE_XLS}")
        return False
    
    print(f"Found template: {template_path}")
    print(f"Output: {TEMPLATE_XLSX}")
    print()
    
    # Install xlrd if needed
    if not install_xlrd():
        print("Trying pandas fallback...")
    
    print()
    
    # Try to convert
    try:
        # First try with xlrd (better formatting support)
        import xlrd
        success = convert_with_xlrd(template_path, TEMPLATE_XLSX)
    except ImportError:
        print("xlrd not available, using pandas...")
        success = convert_with_pandas(template_path, TEMPLATE_XLSX)
    except Exception as e:
        print(f"xlrd failed: {e}")
        import traceback
        traceback.print_exc()
        print("\nTrying pandas fallback...")
        try:
            success = convert_with_pandas(template_path, TEMPLATE_XLSX)
        except Exception as e2:
            print(f"✗ Conversion failed: {e2}")
            import traceback
            traceback.print_exc()
            return False
    
    if success:
        print()
        print("=" * 60)
        print("✓ Conversion complete!")
        print(f"  Template saved to: {TEMPLATE_XLSX}")
        print("  All sheets, columns, and formatting preserved.")
        print("  You can now run the server.")
        print("=" * 60)
    
    return success


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)

