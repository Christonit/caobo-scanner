"""
Test script to verify Excel template population with dummy data
"""
import sys
from pathlib import Path
import openpyxl

# Add parent directory to path
BASE_DIR = Path(__file__).parent
TEMPLATE_FILE = BASE_DIR.parent / "template.xls"
OUTPUT_FILE = BASE_DIR / "test_output.xlsx"

def test_excel_population():
    """Test populating Excel with dummy data"""
    print("Testing Excel template population...")
    
    # Dummy data
    dummy_data = {
        "date": "2024-01-15",
        "vendor": "Starbucks Coffee",
        "total": 12.50,
        "tax": 1.25,
        "filename": "test_receipt.jpg"
    }
    
    # Load template or create new
    if TEMPLATE_FILE.exists():
        print(f"Loading template from: {TEMPLATE_FILE}")
        try:
            wb = openpyxl.load_workbook(TEMPLATE_FILE)
            ws = wb.active
            print(f"Template loaded. Current rows: {ws.max_row}")
        except Exception as e:
            print(f"Error loading template: {e}")
            print("Creating new workbook...")
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Date", "Vendor", "Total", "Tax", "Filename"])
    else:
        print("Template not found. Creating new workbook...")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Date", "Vendor", "Total", "Tax", "Filename"])
    
    # Find first empty row
    row = 1
    while ws.cell(row=row, column=1).value is not None:
        row += 1
    
    print(f"Adding data to row {row}")
    
    # Populate with dummy data
    ws.cell(row=row, column=1, value=dummy_data["date"])
    ws.cell(row=row, column=2, value=dummy_data["vendor"])
    ws.cell(row=row, column=3, value=dummy_data["total"])
    ws.cell(row=row, column=4, value=dummy_data["tax"])
    ws.cell(row=row, column=5, value=dummy_data["filename"])
    
    # Save to output file
    wb.save(OUTPUT_FILE)
    print(f"✓ Test output saved to: {OUTPUT_FILE}")
    print("Excel population test completed successfully!")

if __name__ == "__main__":
    test_excel_population()

