"""
Script to analyze template.xls and generate instructions for server.py
This will read the template and extract all formatting, structure, filters, etc.
"""
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json

BASE_DIR = Path(__file__).parent
TEMPLATE_FILE = BASE_DIR.parent / "template.xls"
OUTPUT_ANALYSIS = BASE_DIR / "template_analysis.json"

def analyze_template():
    """Analyze the template file and extract all structure information"""
    print("=" * 80)
    print("TEMPLATE ANALYSIS SCRIPT")
    print("=" * 80)
    
    if not TEMPLATE_FILE.exists():
        print(f"ERROR: Template file not found: {TEMPLATE_FILE}")
        return None
    
    print(f"\nReading template: {TEMPLATE_FILE}")
    
    # First, convert to xlsx for detailed analysis
    temp_xlsx = BASE_DIR / "temp_template_analysis.xlsx"
    
    try:
        # Read with pandas to get basic structure
        print("\n1. Reading with pandas to get sheet structure...")
        xls_file = pd.ExcelFile(TEMPLATE_FILE, engine='xlrd')
        sheet_names = xls_file.sheet_names
        print(f"   Found {len(sheet_names)} sheet(s): {sheet_names}")
        
        # Convert to xlsx for openpyxl analysis
        print("\n2. Converting to xlsx for detailed formatting analysis...")
        wb_output = openpyxl.Workbook()
        if len(sheet_names) > 0:
            wb_output.remove(wb_output.active)
        
        analysis = {
            "template_path": str(TEMPLATE_FILE),
            "sheets": {},
            "column_mapping": {},
            "formatting": {},
            "filters": {},
            "instructions": []
        }
        
        for sheet_name in sheet_names:
            print(f"\n   Processing sheet: '{sheet_name}'")
            df = pd.read_excel(TEMPLATE_FILE, sheet_name=sheet_name, engine='xlrd', header=None)
            
            # Create worksheet in output
            ws_output = wb_output.create_sheet(title=sheet_name)
            
            # Write data
            for row_idx, row in df.iterrows():
                for col_idx, value in enumerate(row, start=1):
                    if pd.notna(value):
                        ws_output.cell(row=row_idx + 1, column=col_idx, value=value)
            
            # Now analyze with openpyxl
            sheet_analysis = analyze_sheet_with_openpyxl(ws_output, sheet_name)
            analysis["sheets"][sheet_name] = sheet_analysis
        
        # Save converted file for reference
        wb_output.save(temp_xlsx)
        print(f"\n   Converted template saved to: {temp_xlsx}")
        
        # Now load the converted file with openpyxl to get full formatting
        print("\n3. Analyzing formatting with openpyxl...")
        wb = openpyxl.load_workbook(temp_xlsx)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n   Analyzing sheet: '{sheet_name}'")
            
            # Get detailed formatting
            detailed_analysis = get_detailed_formatting(ws, sheet_name)
            analysis["sheets"][sheet_name].update(detailed_analysis)
        
        # Generate column mapping instructions
        print("\n4. Generating column mapping...")
        main_sheet = analysis["sheets"].get("Listado de Gastos", {})
        headers = main_sheet.get("headers", {})
        
        # Map our data fields to template columns
        column_mapping = generate_column_mapping(headers)
        analysis["column_mapping"] = column_mapping
        
        # Generate code instructions
        print("\n5. Generating code instructions...")
        instructions = generate_instructions(analysis)
        analysis["instructions"] = instructions
        
        # Save analysis
        with open(OUTPUT_ANALYSIS, 'w', encoding='utf-8') as f:
            json.dump(analysis, f, indent=2, ensure_ascii=False, default=str)
        
        print(f"\n✓ Analysis saved to: {OUTPUT_ANALYSIS}")
        
        # Print summary
        print("\n" + "=" * 80)
        print("ANALYSIS SUMMARY")
        print("=" * 80)
        print(f"Template: {TEMPLATE_FILE}")
        print(f"Sheets: {', '.join(analysis['sheets'].keys())}")
        
        if "Listado de Gastos" in analysis["sheets"]:
            main = analysis["sheets"]["Listado de Gastos"]
            print(f"\nMain sheet 'Listado de Gastos':")
            print(f"  Headers found: {len(main.get('headers', {}))}")
            if main.get('headers'):
                print("  Column headers:")
                for col_idx, header in sorted(main['headers'].items()):
                    print(f"    Column {col_idx} ({get_column_letter(col_idx)}): '{header}'")
            
            print(f"\n  Column mapping for our data:")
            for field, col_info in analysis['column_mapping'].items():
                print(f"    {field} -> Column {col_info['column']} ({get_column_letter(col_info['column'])})")
        
        print("\n" + "=" * 80)
        print("INSTRUCTIONS FOR server.py")
        print("=" * 80)
        for i, instruction in enumerate(instructions, 1):
            print(f"\n{i}. {instruction}")
        
        return analysis
        
    except Exception as e:
        print(f"\nERROR: {type(e).__name__}: {str(e)}")
        import traceback
        print(traceback.format_exc())
        return None


def analyze_sheet_with_openpyxl(ws, sheet_name):
    """Analyze a worksheet to extract structure"""
    analysis = {
        "name": sheet_name,
        "max_row": ws.max_row,
        "max_column": ws.max_column,
        "headers": {},
        "data_start_row": None,
        "frozen_panes": ws.freeze_panes,
    }
    
    # Find header row (usually row 1, but check for yellow background or bold)
    header_row = 1
    for row_idx in range(1, min(ws.max_row + 1, 5)):
        cell = ws.cell(row=row_idx, column=1)
        if cell.value and isinstance(cell.value, str):
            # Check if this looks like a header row
            fill = cell.fill
            font = cell.font
            if (fill and fill.start_color and fill.start_color.rgb and 
                'FFFF' in str(fill.start_color.rgb).upper()) or font.bold:
                header_row = row_idx
                analysis["data_start_row"] = row_idx + 1
                break
    
    # Extract headers
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        if cell.value:
            analysis["headers"][col_idx] = str(cell.value).strip()
    
    if not analysis["data_start_row"]:
        analysis["data_start_row"] = header_row + 1
    
    return analysis


def get_detailed_formatting(ws, sheet_name):
    """Get detailed formatting information from a worksheet"""
    formatting = {
        "header_row": None,
        "header_formatting": {},
        "column_widths": {},
        "row_heights": {},
        "cell_formats": {},
        "filters": {},
    }
    
    # Find header row
    header_row = 1
    for row_idx in range(1, min(ws.max_row + 1, 5)):
        cell = ws.cell(row=row_idx, column=1)
        if cell.value:
            fill = cell.fill
            if fill and fill.start_color:
                # Check for yellow (common header color)
                rgb = str(fill.start_color.rgb or '').upper()
                if 'FFFF' in rgb or 'FFEB' in rgb:
                    header_row = row_idx
                    break
    
    formatting["header_row"] = header_row
    
    # Analyze header row formatting
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        if cell.value:
            cell_format = {
                "font": {
                    "name": cell.font.name,
                    "size": cell.font.size,
                    "bold": cell.font.bold,
                    "color": str(cell.font.color.rgb) if cell.font.color and cell.font.color.rgb else None,
                },
                "fill": {
                    "patternType": cell.fill.patternType if cell.fill else None,
                    "fgColor": str(cell.fill.fgColor.rgb) if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb else None,
                },
                "alignment": {
                    "horizontal": cell.alignment.horizontal if cell.alignment else None,
                    "vertical": cell.alignment.vertical if cell.alignment else None,
                },
            }
            formatting["header_formatting"][col_idx] = cell_format
    
    # Get column widths
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        width = ws.column_dimensions[col_letter].width
        if width:
            formatting["column_widths"][col_idx] = width
    
    # Check for auto filters
    if ws.auto_filter.ref:
        formatting["filters"]["enabled"] = True
        formatting["filters"]["range"] = str(ws.auto_filter.ref)
    else:
        formatting["filters"]["enabled"] = False
    
    return formatting


def generate_column_mapping(headers):
    """Generate mapping from our data fields to template columns"""
    mapping = {}
    
    # Our data fields
    our_fields = {
        "date": ["fecha", "date", "fecha de", "fecha del"],
        "vendor": ["proveedor", "vendor", "suplidor", "supplier"],
        "description": ["descripcion", "description", "desc"],
        "total": ["monto", "total", "amount", "monto en servicios"],
        "tax": ["impuesto", "tax", "iva"],
        "filename": ["archivo", "filename", "file", "documento"],
    }
    
    # Find matching columns
    for field, keywords in our_fields.items():
        for col_idx, header_text in headers.items():
            header_lower = header_text.lower()
            if any(keyword in header_lower for keyword in keywords):
                mapping[field] = {
                    "column": col_idx,
                    "header": header_text,
                    "column_letter": get_column_letter(col_idx),
                }
                break
    
    return mapping


def generate_instructions(analysis):
    """Generate code instructions for server.py"""
    instructions = []
    
    main_sheet_name = "Listado de Gastos"
    if main_sheet_name not in analysis["sheets"]:
        instructions.append("ERROR: Main sheet 'Listado de Gastos' not found in template")
        return instructions
    
    main_sheet = analysis["sheets"][main_sheet_name]
    column_mapping = analysis["column_mapping"]
    formatting = main_sheet.get("formatting", {})
    
    instructions.append(f"Always load template from: {analysis['template_path']}")
    instructions.append(f"Use sheet name: '{main_sheet_name}' (not active sheet)")
    instructions.append(f"Data should start at row: {main_sheet.get('data_start_row', 2)}")
    
    if column_mapping:
        instructions.append("\nColumn mapping (our data -> template column):")
        for field, col_info in column_mapping.items():
            instructions.append(f"  {field} -> Column {col_info['column']} ({col_info['column_letter']}): '{col_info['header']}'")
    
    if formatting.get("header_formatting"):
        instructions.append("\nHeader row formatting:")
        instructions.append(f"  Header row: {formatting.get('header_row', 1)}")
        instructions.append("  Apply yellow background and bold font to header row")
    
    if formatting.get("filters", {}).get("enabled"):
        instructions.append(f"\nAuto-filter is enabled in template")
        instructions.append("  Apply auto-filter to header row after writing data")
    
    if formatting.get("column_widths"):
        instructions.append("\nColumn widths to preserve:")
        for col_idx, width in formatting["column_widths"].items():
            instructions.append(f"  Column {get_column_letter(col_idx)}: {width}")
    
    instructions.append("\nCode structure:")
    instructions.append("  1. Always start from template (never from output file)")
    instructions.append("  2. Load the specific sheet 'Listado de Gastos'")
    instructions.append("  3. Find first empty row after header")
    instructions.append("  4. Map data fields to template columns using the mapping above")
    instructions.append("  5. Apply header formatting if not already present")
    instructions.append("  6. Apply auto-filter to header row")
    instructions.append("  7. Preserve all column widths")
    instructions.append("  8. Save to output file")
    
    return instructions


if __name__ == "__main__":
    analyze_template()

